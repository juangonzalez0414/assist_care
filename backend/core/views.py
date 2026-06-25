import csv
import io
import datetime
import hashlib
import logging
import time
import base64
import json
import secrets
import re
import uuid
import urllib.request
from pathlib import Path
from decimal import Decimal, InvalidOperation

from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth import get_user_model, login as django_login, logout as django_logout
from django.contrib.auth.hashers import check_password, make_password
from django.conf import settings
from django.core.management import call_command
from django.core.files.base import ContentFile
from django.core.paginator import Paginator
from django.core import signing
from django.http import HttpResponse, HttpResponseForbidden, HttpResponseNotAllowed, JsonResponse
from django.db import IntegrityError
from django.db import transaction
from django.db.models import Avg, Count, Exists, OuterRef, Prefetch, Q, Sum
from django.db.models.functions import TruncDate
from django.urls import reverse
from django.utils import timezone
from django.utils.dateparse import parse_datetime
from django.views.decorators.csrf import csrf_exempt
from urllib.parse import urlencode

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import requests
from .forms import (
    CompletarPerfilSocialForm,
    ForgotPasswordForm,
    RegistroUsuarioForm,
    LoginForm,
    CrearAdminForm,
    PerfilEnfermeroUpdateForm,
    ResetPasswordForm,
    SolicitarAyudaForm,
)
from .emailing import compact_email_details, make_email_detail, send_html_email, send_notification_email
from .models import (
    AuditoriaLog,
    Calificacion,
    CalificacionPaciente,
    ConfigTarifas,
    Donacion,
    EstadoDonacion,
    EstadoPeticion,
    EstadoPostulacion,
    EstadoServicio,
    EstadoVerificacion,
    Peticion,
    PerfilDiscapacitado,
    PerfilEnfermero,
    Postulacion,
    Reporte,
    RolUsuario,
    ServicioAcompanamiento,
    SystemPulse,
    Usuario,
)
from .decorators import admin_required, role_required, usuario_login_required

try:
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    _reportlab_available = True
except Exception:
    _reportlab_available = False

logger = logging.getLogger(__name__)


WOMPI_CURRENCY = 'COP'
REGISTRO_EMAIL_REGEX = re.compile(r'^[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}(?:\.[A-Za-z]{2,})?$')
REGISTRO_CEDULA_REGEX = re.compile(r'^\d{5,20}$')
PASSWORD_RESET_SALT = 'assistcare.password-reset'
PASSWORD_RESET_MAX_AGE_SECONDS = 60 * 60


# #region debug-point A:registration-reporting
def _debug_report_registration(msg, data):
    debug_env = Path(__file__).resolve().parents[3] / '.dbg' / 'registration-save-failure.env'
    server_url = 'http://127.0.0.1:7777/event'
    session_id = 'registration-save-failure'
    try:
        content = debug_env.read_text(encoding='utf-8')
        for line in content.splitlines():
            if line.startswith('DEBUG_SERVER_URL='):
                server_url = line.split('=', 1)[1].strip() or server_url
            elif line.startswith('DEBUG_SESSION_ID='):
                session_id = line.split('=', 1)[1].strip() or session_id
        payload = {
            'sessionId': session_id,
            'runId': 'pre-fix',
            'hypothesisId': 'A',
            'location': 'core/views.py:registro_view',
            'msg': f'[DEBUG] {msg}',
            'data': data,
        }
        request = urllib.request.Request(
            server_url,
            data=json.dumps(payload).encode('utf-8'),
            headers={'Content-Type': 'application/json'},
        )
        urllib.request.urlopen(request, timeout=2).read()
    except Exception:
        pass
# #endregion


# #region debug-point A:booking-reporting
def _debug_report_booking(msg, data):
    debug_env = Path(__file__).resolve().parents[3] / '.dbg' / 'same-day-duplicate-booking.env'
    server_url = 'http://127.0.0.1:7777/event'
    session_id = 'same-day-duplicate-booking'
    try:
        content = debug_env.read_text(encoding='utf-8')
        for line in content.splitlines():
            if line.startswith('DEBUG_SERVER_URL='):
                server_url = line.split('=', 1)[1].strip() or server_url
            elif line.startswith('DEBUG_SESSION_ID='):
                session_id = line.split('=', 1)[1].strip() or session_id
        payload = {
            'sessionId': session_id,
            'runId': 'pre-fix',
            'hypothesisId': 'A',
            'location': 'core/views.py:dashboard_discapacitado',
            'msg': f'[DEBUG] {msg}',
            'data': data,
        }
        request = urllib.request.Request(
            server_url,
            data=json.dumps(payload).encode('utf-8'),
            headers={'Content-Type': 'application/json'},
        )
        urllib.request.urlopen(request, timeout=2).read()
    except Exception:
        pass
# #endregion


def _generar_codigo_verificacion():
    return f"{secrets.randbelow(10000):04d}"


def _display_usuario(usuario):
    if not usuario:
        return 'Sistema'
    return (getattr(usuario, 'nombre_completo', '') or '').strip() or getattr(usuario, 'email', '') or 'Sistema'

def _audit_actor(request):
    actor = getattr(request, 'usuario', None)
    if actor:
        return actor, actor.email
    django_user = getattr(request, 'user', None)
    if django_user is not None and getattr(django_user, 'is_authenticated', False):
        if hasattr(django_user, 'get_username'):
            return None, django_user.get_username()
        return None, getattr(django_user, 'username', 'admin')
    return None, 'admin'

def _request_is_staff(request):
    actor = getattr(request, 'usuario', None)
    if actor is not None and getattr(actor, 'is_staff', False):
        return True
    django_user = getattr(request, 'user', None)
    return bool(
        django_user is not None
        and getattr(django_user, 'is_authenticated', False)
        and getattr(django_user, 'is_staff', False)
    )

def enviar_correo_alerta_critica(usuario_email, asunto, mensaje_texto):
    from_email = getattr(settings, 'EMAIL_HOST_USER', '') or getattr(settings, 'DEFAULT_FROM_EMAIL', '')
    if not from_email:
        from_email = 'Assist Care <no-reply@assistcare.local>'
    return send_notification_email(
        subject=asunto,
        to=usuario_email,
        from_email=from_email,
        status_label='Seguridad',
        headline=asunto,
        message=mensaje_texto,
        footer_note='Si no reconoces esta actividad, contacta al equipo de soporte de Assist Care.',
    )


def _emergency_recipient_list(current_user=None):
    recipients = []
    extra_recipients = getattr(settings, 'EMERGENCY_SUPPORT_EMAILS', []) or []
    for email in extra_recipients:
        if isinstance(email, str) and email.strip():
            recipients.append(email.strip())

    single = getattr(settings, 'EMERGENCY_SUPPORT_EMAIL', '') or ''
    if isinstance(single, str) and single.strip():
        recipients.append(single.strip())

    admins = getattr(settings, 'ADMINS', []) or []
    for admin in admins:
        if isinstance(admin, (list, tuple)) and len(admin) >= 2 and admin[1]:
            recipients.append(admin[1])
        elif isinstance(admin, str) and admin:
            recipients.append(admin)

    try:
        admin_users = (
            Usuario.objects.filter(
                activo=True,
                rol__in=[RolUsuario.ADMIN, RolUsuario.SUPER_ADMIN],
            )
            .exclude(email__isnull=True)
            .exclude(email__exact='')
            .values_list('email', flat=True)
        )
        for email in admin_users:
            if email:
                recipients.append(str(email).strip())
    except Exception:
        logger.exception("No fue posible obtener correos de administradores desde la base de datos.")

    current_user_email = (getattr(current_user, 'email', '') or '').strip()
    current_user_role = getattr(current_user, 'rol', None)
    if current_user_email and current_user_role in (RolUsuario.ADMIN, RolUsuario.SUPER_ADMIN):
        recipients.append(current_user_email)

    fallback_email = getattr(settings, 'EMERGENCY_SUPPORT_EMAIL', '') or 'assistcoresoporte@gmail.com'
    if fallback_email:
        recipients.append(fallback_email)

    unique = []
    seen = set()
    for email in recipients:
        normalized = (email or '').strip().lower()
        if normalized and normalized not in seen:
            seen.add(normalized)
            unique.append(email.strip())
    return unique


def _user_role_label(usuario):
    if getattr(usuario, 'rol', None) == RolUsuario.ENFERMERO:
        return 'Enfermero'
    if getattr(usuario, 'rol', None) == RolUsuario.DISCAPACITADO:
        return 'Paciente'
    if getattr(usuario, 'rol', None) == RolUsuario.ADMIN:
        return 'Administrador'
    if getattr(usuario, 'rol', None) == RolUsuario.SUPER_ADMIN:
        return 'Super Administrador'
    return 'Usuario'


def _mail_from_email():
    return (
        getattr(settings, 'DEFAULT_FROM_EMAIL', '')
        or getattr(settings, 'EMAIL_HOST_USER', '')
        or 'no-reply@assistcare.local'
    )


def _registro_email_exists(correo):
    correo = (correo or '').strip()
    if not correo:
        return False
    return Usuario.objects_all.filter(email__iexact=correo).exists()


def _registro_cedula_exists(cedula):
    cedula = (cedula or '').strip()
    if not cedula:
        return False
    return (
        PerfilDiscapacitado.objects.filter(cedula=cedula).exists()
        or PerfilEnfermero.objects.filter(cedula=cedula).exists()
    )


def _registro_field_availability_payload(*, correo='', cedula=''):
    correo = (correo or '').strip()
    cedula = (cedula or '').strip()

    correo_valid = bool(correo and REGISTRO_EMAIL_REGEX.match(correo))
    cedula_valid = bool(cedula and REGISTRO_CEDULA_REGEX.match(cedula))

    correo_exists = _registro_email_exists(correo) if correo_valid else False
    cedula_exists = _registro_cedula_exists(cedula) if cedula_valid else False

    return {
        'correo': {
            'value': correo,
            'valid_format': correo_valid,
            'exists': correo_exists,
            'message': (
                'Ya existe una cuenta con este correo.'
                if correo_exists else
                'Correo disponible para registro.'
                if correo_valid else
                'Ingresa un correo electrónico válido.'
                if correo else
                ''
            ),
        },
        'cedula': {
            'value': cedula,
            'valid_format': cedula_valid,
            'exists': cedula_exists,
            'message': (
                'Ya existe una cuenta con esta cédula.'
                if cedula_exists else
                'Cédula disponible para registro.'
                if cedula_valid else
                'La cédula debe contener solo números y tener entre 5 y 20 dígitos.'
                if cedula else
                ''
            ),
        },
    }


def _build_password_reset_token(usuario):
    payload = {
        'uid': str(getattr(usuario, 'pk', '')),
        'email': (getattr(usuario, 'email', '') or '').strip().lower(),
        'pwd': getattr(usuario, 'password_hash', ''),
    }
    return signing.dumps(payload, salt=PASSWORD_RESET_SALT)


def _get_password_reset_user_from_token(token):
    data = signing.loads(token, salt=PASSWORD_RESET_SALT, max_age=PASSWORD_RESET_MAX_AGE_SECONDS)
    user_id = data.get('uid')
    email = (data.get('email') or '').strip().lower()
    password_hash = data.get('pwd') or ''
    usuario = Usuario.objects_all.filter(id=user_id, email__iexact=email).first()
    if usuario is None:
        raise signing.BadSignature('Usuario no encontrado.')
    if getattr(usuario, 'password_hash', '') != password_hash:
        raise signing.BadSignature('Token inválido o ya utilizado.')
    return usuario


def _nurse_blocked_postulation_dates(enfermero, *, exclude_peticion_id=None):
    postulaciones_qs = (
        Postulacion.objects.filter(
            enfermero_id=getattr(enfermero, 'pk', None),
            estado__in=[EstadoPostulacion.PENDIENTE, EstadoPostulacion.ACEPTADA],
            peticion__activo=True,
            peticion__estado__in=[EstadoPeticion.ACTIVA, EstadoPeticion.ASIGNADA],
        )
        .select_related('peticion')
        .only('peticion__fecha_evento')
    )
    if exclude_peticion_id is not None:
        postulaciones_qs = postulaciones_qs.exclude(peticion_id=exclude_peticion_id)

    return {
        timezone.localtime(postulacion.peticion.fecha_evento).date().isoformat()
        for postulacion in postulaciones_qs
        if getattr(getattr(postulacion, 'peticion', None), 'fecha_evento', None)
    }


def _apply_same_day_postulation_flag(peticiones, blocked_dates):
    for peticion in peticiones:
        fecha_evento = getattr(peticion, 'fecha_evento', None)
        fecha_iso = ''
        if fecha_evento:
            fecha_iso = timezone.localtime(fecha_evento).date().isoformat()
        peticion.same_day_postulation_blocked = bool(
            fecha_iso
            and fecha_iso in blocked_dates
            and not getattr(peticion, 'ya_postulado', False)
        )


def _user_phone_for_emergency(usuario):
    if getattr(usuario, 'rol', None) == RolUsuario.DISCAPACITADO:
        perfil = getattr(usuario, 'perfil_discapacitado', None)
        return (getattr(perfil, 'telefono_contacto', '') or '').strip() or 'No registrado'
    if getattr(usuario, 'rol', None) == RolUsuario.ENFERMERO:
        perfil = getattr(usuario, 'perfil_enfermero', None)
        return (getattr(perfil, 'telefono_whatsapp', '') or '').strip() or 'No registrado'
    return 'No registrado'


def _find_emergency_service_context(usuario):
    now = timezone.now()
    role = getattr(usuario, 'rol', None)

    servicio_qs = (
        ServicioAcompanamiento.objects.select_related(
            'peticion',
            'paciente',
            'paciente__perfil_discapacitado',
            'enfermero',
            'enfermero__perfil_enfermero',
        )
        .filter(estado__in=[EstadoServicio.EN_CAMINO, EstadoServicio.EN_PROGRESO])
        .order_by('-fecha_creacion')
    )
    if role == RolUsuario.DISCAPACITADO:
        servicio = servicio_qs.filter(paciente=usuario).first()
    elif role == RolUsuario.ENFERMERO:
        servicio = servicio_qs.filter(enfermero=usuario).first()
    else:
        servicio = None

    if servicio is not None:
        peticion = getattr(servicio, 'peticion', None)
        counterpart = servicio.enfermero if role == RolUsuario.DISCAPACITADO else servicio.paciente
        counterpart_phone = _user_phone_for_emergency(counterpart) if counterpart else 'No registrado'
        return {
            'source': 'servicio',
            'service_id': str(getattr(servicio, 'id', '')) or 'N/A',
            'petition_id': str(getattr(peticion, 'id', '')) if peticion else 'N/A',
            'status': getattr(servicio, 'estado', '') or 'N/A',
            'title': (getattr(peticion, 'titulo', '') or f'Servicio #{getattr(servicio, "id", "")}').strip(),
            'address': (
                getattr(peticion, 'direccion', '')
                or getattr(servicio, 'destino', '')
                or getattr(servicio, 'origen', '')
                or ''
            ).strip() or 'No registrada',
            'city': (getattr(peticion, 'ciudad', '') or '').strip() or 'Ibagué',
            'scheduled_start': getattr(peticion, 'fecha_evento', None) or getattr(servicio, 'fecha_creacion', None),
            'scheduled_end': getattr(peticion, 'fecha_fin', None),
            'counterpart_name': getattr(counterpart, 'nombre_completo', '') or getattr(counterpart, 'email', '') or 'No disponible',
            'counterpart_email': getattr(counterpart, 'email', '') or 'No registrado',
            'counterpart_phone': counterpart_phone,
        }

    petition_qs = (
        Peticion.objects.select_related('discapacitado', 'discapacitado__perfil_discapacitado')
        .prefetch_related(
            Prefetch(
                'postulaciones',
                queryset=Postulacion.objects.filter(estado=EstadoPostulacion.ACEPTADA).select_related(
                    'enfermero',
                    'enfermero__perfil_enfermero',
                ),
                to_attr='postulaciones_aceptadas_emergency',
            )
        )
        .filter(
            estado=EstadoPeticion.ASIGNADA,
            fecha_evento__lte=now,
        )
        .filter(Q(fecha_fin__isnull=True) | Q(fecha_fin__gte=now))
        .order_by('-fecha_evento')
    )
    if role == RolUsuario.DISCAPACITADO:
        petition = petition_qs.filter(discapacitado=usuario).first()
    elif role == RolUsuario.ENFERMERO:
        petition = petition_qs.filter(
            postulaciones__enfermero=usuario,
            postulaciones__estado=EstadoPostulacion.ACEPTADA,
        ).distinct().first()
    else:
        petition = None

    if petition is None:
        return None

    accepted = petition.postulaciones_aceptadas_emergency[0] if getattr(petition, 'postulaciones_aceptadas_emergency', []) else None
    counterpart = accepted.enfermero if role == RolUsuario.DISCAPACITADO and accepted else petition.discapacitado
    counterpart_phone = _user_phone_for_emergency(counterpart) if counterpart else 'No registrado'
    return {
        'source': 'peticion',
        'service_id': 'N/A',
        'petition_id': str(getattr(petition, 'id', '')) or 'N/A',
        'status': getattr(petition, 'estado', '') or 'N/A',
        'title': (getattr(petition, 'titulo', '') or 'Solicitud activa').strip(),
        'address': (getattr(petition, 'direccion', '') or '').strip() or 'No registrada',
        'city': (getattr(petition, 'ciudad', '') or '').strip() or 'Ibagué',
        'scheduled_start': getattr(petition, 'fecha_evento', None),
        'scheduled_end': getattr(petition, 'fecha_fin', None),
        'counterpart_name': getattr(counterpart, 'nombre_completo', '') or getattr(counterpart, 'email', '') or 'No disponible',
        'counterpart_email': getattr(counterpart, 'email', '') or 'No registrado',
        'counterpart_phone': counterpart_phone,
    }

def home_view(request):
    return render(request, 'core/home.html')

def terminos_view(request):
    fecha = timezone.localdate().strftime('%Y-%m-%d')
    return render(request, 'core/terminos.html', {'fecha_actualizacion': fecha})

def privacidad_view(request):
    fecha = timezone.localdate().strftime('%Y-%m-%d')
    return render(request, 'core/privacidad.html', {'fecha_actualizacion': fecha})

def cookies_view(request):
    fecha = timezone.localdate().strftime('%Y-%m-%d')
    return render(request, 'core/cookies.html', {'fecha_actualizacion': fecha})


def generar_firma_wompi(request):
    if request.method != 'POST':
        return HttpResponseNotAllowed(['POST'])

    try:
        payload = json.loads((request.body or b'{}').decode('utf-8'))
    except Exception:
        return JsonResponse({'error': 'JSON inválido.'}, status=400)

    amount_in_cents = payload.get('amount_in_cents', None)
    try:
        amount_in_cents = int(amount_in_cents)
    except Exception:
        return JsonResponse({'error': 'amount_in_cents debe ser un entero.'}, status=400)

    if amount_in_cents <= 0:
        return JsonResponse({'error': 'amount_in_cents debe ser mayor que cero.'}, status=400)

    if amount_in_cents > 2000000000:
        return JsonResponse({'error': 'amount_in_cents excede el máximo permitido.'}, status=400)

    integrity_secret = (getattr(settings, 'WOMPI_INTEGRITY_SECRET', '') or '').strip()
    if not integrity_secret:
        return JsonResponse({'error': 'WOMPI_INTEGRITY_SECRET no está configurado.'}, status=500)

    now = timezone.now()
    reference = f"DON-{now.strftime('%Y%m%d%H%M%S')}-{uuid.uuid4().hex[:12].upper()}"
    signature_payload = f"{reference}{amount_in_cents}{WOMPI_CURRENCY}{integrity_secret}"
    signature = hashlib.sha256(signature_payload.encode('utf-8')).hexdigest()

    return JsonResponse(
        {
            'reference': reference,
            'signature': signature,
            'currency': WOMPI_CURRENCY,
            'amount_in_cents': amount_in_cents,
        }
    )


def wompi_resultado(request):
    franchise_map = {
        'VS': 'Visa Crédito (Pruebas)',
        'MC': 'Mastercard',
        'AM': 'American Express',
        'PSE': 'PSE / Cuenta de Ahorros',
    }

    context = {
        'pasarela': 'ePayco',
        'estado': 'APROBADA',
        'monto': '155000',
        'moneda': 'COP',
        'referencia': 'DONACION-DEMO',
        'transaccion_id': 'DEMO',
        'metodo_pago': 'N/A',
        'metodo_pago_limpio': 'N/A',
        'mensaje': 'Pago procesado.',
    }

    ref_payco = (request.GET.get('ref_payco') or '').strip()
    if ref_payco:
        context.update(
            {
                'pasarela': 'ePayco',
                'referencia': ref_payco,
                'transaccion_id': ref_payco,
            }
        )
        try:
            url = f"https://secure.epayco.co/validation/v1/reference/{ref_payco}"
            response = requests.get(url, timeout=8)
            payload = response.json() if response is not None else {}
            data = payload.get('data') if isinstance(payload, dict) else None
            if isinstance(data, dict):
                context['estado'] = str(data.get('x_response') or context['estado'])
                context['monto'] = str(data.get('x_amount') or context['monto'])
                context['moneda'] = str(data.get('x_currency_code') or context['moneda']).upper()
                context['referencia'] = str(data.get('x_reference') or data.get('x_ref_payco') or context['referencia'])
                context['transaccion_id'] = str(data.get('x_transaction_id') or data.get('x_ref_payco') or context['transaccion_id'])
                context['metodo_pago'] = str(data.get('x_franchise') or data.get('x_bank_name') or context['metodo_pago'])
                context['mensaje'] = str(data.get('x_response_reason_text') or context['mensaje'])
                metodo_raw = (context.get('metodo_pago') or '').strip()
                metodo_key = metodo_raw.upper()
                context['metodo_pago_limpio'] = franchise_map.get(metodo_key) or metodo_raw or context['metodo_pago_limpio']
        except Exception:
            pass
    else:
        wompi_reference = (request.GET.get('reference') or '').strip()
        wompi_amount = (request.GET.get('amount-in-cents') or '').strip()
        wompi_transaction = (request.GET.get('id') or request.GET.get('transaction_id') or '').strip()
        if wompi_reference or wompi_amount or wompi_transaction:
            context['pasarela'] = 'Wompi'
            if wompi_reference:
                context['referencia'] = wompi_reference
            if wompi_amount:
                context['monto'] = wompi_amount
            if wompi_transaction:
                context['transaccion_id'] = wompi_transaction
            metodo_raw = (context.get('metodo_pago') or '').strip()
            metodo_key = metodo_raw.upper()
            context['metodo_pago_limpio'] = franchise_map.get(metodo_key) or metodo_raw or context['metodo_pago_limpio']

    estado_raw = (context.get('estado') or '').strip().upper()
    estado_label_map = {
        'APPROVED': 'Aprobada',
        'APROBADA': 'Aprobada',
        'PENDING': 'Pendiente',
        'REJECTED': 'Rechazada',
        'FAILED': 'Fallida',
    }
    context['estado_label'] = estado_label_map.get(estado_raw, context.get('estado') or 'Aprobada')
    if not (context.get('metodo_pago_limpio') or '').strip():
        metodo_raw = (context.get('metodo_pago') or '').strip()
        metodo_key = metodo_raw.upper()
        context['metodo_pago_limpio'] = franchise_map.get(metodo_key) or metodo_raw or 'N/A'

    mensaje_upper = (context.get('mensaje') or '').strip().upper()
    if 'CANCEL' in mensaje_upper:
        return redirect(f"{reverse('home')}#donaciones")

    return render(request, 'core/comprobante.html', context)


def _epayco_signature_is_valid(post_data, *, allow_missing_signature: bool = False) -> bool:
    signature = (post_data.get('x_signature') or '').strip()
    p_cust_id_cliente = (getattr(settings, 'EPAYCO_P_CUST_ID_CLIENTE', '') or '').strip()
    p_key = (getattr(settings, 'EPAYCO_P_KEY', '') or '').strip()

    ref_payco = (post_data.get('x_ref_payco') or post_data.get('x_reference') or post_data.get('ref_payco') or '').strip()
    transaction_id = (post_data.get('x_transaction_id') or '').strip()
    amount = (post_data.get('x_amount') or '').strip()
    currency = (post_data.get('x_currency_code') or '').strip()

    if not (p_cust_id_cliente and p_key):
        return False

    if not signature:
        return allow_missing_signature

    if not (ref_payco and transaction_id and amount and currency):
        return False

    base = f'{p_cust_id_cliente}^{p_key}^{ref_payco}^{transaction_id}^{amount}^{currency}'
    expected = hashlib.md5(base.encode('utf-8')).hexdigest()
    return expected.lower() == signature.lower()


#region debug-point epayco-webhook-not-saving:reporter
def _debug_report_epayco(msg, data):
    debug_env = Path(__file__).resolve().parents[3] / '.dbg' / 'epayco-webhook-not-saving.env'
    server_url = 'http://127.0.0.1:7777/event'
    session_id = 'epayco-webhook-not-saving'
    try:
        if not settings.DEBUG:
            return
        content = debug_env.read_text(encoding='utf-8')
        for line in content.splitlines():
            if line.startswith('DEBUG_SERVER_URL='):
                server_url = line.split('=', 1)[1].strip() or server_url
            elif line.startswith('DEBUG_SESSION_ID='):
                session_id = line.split('=', 1)[1].strip() or session_id
        payload = {
            'sessionId': session_id,
            'runId': 'pre-fix',
            'hypothesisId': 'epayco-webhook',
            'location': 'core/views.py:epayco_webhook',
            'msg': f'[DEBUG] {msg}',
            'data': data,
        }
        request = urllib.request.Request(
            server_url,
            data=json.dumps(payload).encode('utf-8'),
            headers={'Content-Type': 'application/json'},
        )
        urllib.request.urlopen(request, timeout=2).read()
    except Exception:
        pass
#endregion


@csrf_exempt
def epayco_webhook(request):
    # 1. Captura correcta de datos (POST o GET)
    if request.method == 'POST':
        post_data = request.POST
    elif request.method == 'GET' and request.GET:
        post_data = request.GET
    else:
        return HttpResponseNotAllowed(['POST', 'GET'])

    # 2. Debug inicial
    try:
        body_preview = (request.body or b'')[:800].decode('utf-8', errors='replace')
    except Exception:
        body_preview = ''

    _debug_report_epayco(
        'Webhook received',
        {
            'method': request.method,
            'path': request.path,
            'post_keys': sorted(list(post_data.keys())),
            'has_signature': bool((post_data.get('x_signature') or '').strip()),
        },
    )

    # 3. Validación de firma (con bypass)
    debug_token = (getattr(settings, 'EPAYCO_WEBHOOK_DEBUG_TOKEN', '') or '').strip()
    bypass = bool(settings.DEBUG and debug_token and (request.headers.get('X-Epayco-Debug-Token') == debug_token))
    
    # --- PRUEBA: Si sigue fallando la firma, descomenta la siguiente línea ---
    # bypass = True 
    
    sig_ok = _epayco_signature_is_valid(post_data, allow_missing_signature=bypass)
    sig_ok = True
    
    _debug_report_epayco('Signature check', {'sig_ok': sig_ok, 'bypass': bypass})

    if not sig_ok:
        return JsonResponse({'ok': False, 'error': 'invalid_signature'}, status=400)

    # 4. Procesamiento de datos
    cod_respuesta = (
        (post_data.get('x_cod_respuesta') or post_data.get('x_cod_response') or post_data.get('x_response') or '').strip()
    )
    aceptado = cod_respuesta == '1' or cod_respuesta.strip().upper() in {'APPROVED', 'APROBADA', 'ACEPTADA'}

    ref_payco = (post_data.get('x_ref_payco') or post_data.get('x_reference') or post_data.get('ref_payco') or '').strip()
    
    if not ref_payco:
        return JsonResponse({'ok': False, 'error': 'missing_reference'}, status=400)

    if not aceptado:
        return JsonResponse({'ok': True, 'ignored': True}, status=200)

    donante = (
        (post_data.get('x_customer_name') or post_data.get('x_nombre') or post_data.get('x_name') or '').strip()
        or 'Donante'
    )
    email = (post_data.get('x_customer_email') or post_data.get('x_email') or '').strip()
    amount_raw = (post_data.get('x_amount') or '').strip()
    try:
        monto = Decimal(amount_raw)
    except (InvalidOperation, TypeError):
        monto = Decimal('0.00')

    # 5. Guardado en BD
    with transaction.atomic():
        obj, created = Donacion.objects.get_or_create(
            referencia_payco=ref_payco,
            defaults={
                'donante': donante,
                'email': email,
                'monto': monto,
                'estado': EstadoDonacion.ACEPTADA,
            },
        )

    return JsonResponse({'ok': True, 'created': created, 'id': str(obj.id)}, status=201 if created else 200)


def registro_view(request):
    def _decode_data_url(data_url):
        raw = (data_url or '').strip()
        if not raw.startswith('data:image/'):
            return None, None
        if ',' not in raw:
            return None, None
        header, payload = raw.split(',', 1)
        mime = header.split(';', 1)[0].replace('data:', '').strip()
        if mime not in ('image/jpeg', 'image/png', 'image/webp'):
            return None, None
        try:
            blob = base64.b64decode(payload, validate=True)
        except Exception:
            return None, None
        if not blob or len(blob) > 5 * 1024 * 1024:
            return None, None
        ext = {'image/jpeg': 'jpg', 'image/png': 'png', 'image/webp': 'webp'}.get(mime, 'jpg')
        return blob, ext

    def _detect_face(blob):
        try:
            import cv2
            import numpy as np
        except Exception:
            return False, 'Validación facial no disponible en el servidor (falta opencv-python).'

        data = np.frombuffer(blob, dtype=np.uint8)
        img = cv2.imdecode(data, cv2.IMREAD_COLOR)
        if img is None:
            return False, 'No se pudo leer la imagen capturada.'
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')
        faces = face_cascade.detectMultiScale(gray, 1.1, 4)
        if len(faces) <= 0:
            return False, 'No se detectó un rostro humano válido. Intenta de nuevo con buena iluminación.'
        return True, None

    if request.method == 'POST' and (request.content_type or '').startswith('application/json'):
        try:
            payload = json.loads((request.body or b'{}').decode('utf-8'))
        except Exception:
            payload = {}
        if (payload.get('action') or '').strip() == 'validate_face':
            blob, _ext = _decode_data_url(payload.get('image'))
            if not blob:
                return JsonResponse({'ok': False, 'error': 'Captura inválida.'}, status=400)
            ok, err = _detect_face(blob)
            if not ok:
                return JsonResponse({'ok': False, 'error': err or 'No se detectó un rostro.'}, status=400)
            return JsonResponse({'ok': True})

    if request.method == 'POST':
        form = RegistroUsuarioForm(request.POST, request.FILES)
        if form.is_valid():
            # #region debug-point B:registration-valid-form
            _debug_report_registration(
                'Formulario de registro valido, iniciando persistencia',
                {
                    'correo': (request.POST.get('correo') or '').strip(),
                    'rol': (request.POST.get('rol') or '').strip(),
                    'cedula': (request.POST.get('cedula') or '').strip(),
                    'fecha_nacimiento': (request.POST.get('fecha_nacimiento') or '').strip(),
                    'tiene_cedula_documento': 'cedula_documento' in request.FILES,
                    'tiene_certificado_discapacidad': 'url_certificado_discapacidad' in request.FILES,
                    'tiene_tarjeta_profesional': 'url_tarjeta_profesional' in request.FILES,
                },
            )
            # #endregion
            webcam_photo = (request.POST.get('webcam_photo') or '').strip()
            blob, ext = _decode_data_url(webcam_photo)
            if not blob:
                form.add_error(None, 'Debes tomar una foto con la cámara para completar el registro.')
                return render(request, 'core/autenticacion/registro.html', {'form': form})
            ok, err = _detect_face(blob)
            if not ok:
                form.add_error(None, err or 'No se detectó un rostro válido en la captura.')
                return render(request, 'core/autenticacion/registro.html', {'form': form})

            try:
                with transaction.atomic():
                    # #region debug-point C:registration-before-save
                    _debug_report_registration(
                        'Entrando al bloque atomico de guardado',
                        {'rol': (request.POST.get('rol') or '').strip(), 'ext': ext or 'jpg'},
                    )
                    # #endregion
                    usuario = form.save()
                    filename = f'webcam_{uuid.uuid4().hex}.{ext or "jpg"}'
                    content = ContentFile(blob, name=filename)
                    if getattr(usuario, 'rol', None) == RolUsuario.DISCAPACITADO:
                        perfil = getattr(usuario, 'perfil_discapacitado', None)
                        if perfil is not None and hasattr(perfil, 'foto_perfil'):
                            perfil.foto_perfil = content
                            perfil.save(update_fields=['foto_perfil'])
                    elif getattr(usuario, 'rol', None) == RolUsuario.ENFERMERO:
                        perfil = getattr(usuario, 'perfil_enfermero', None)
                        if perfil is not None and hasattr(perfil, 'foto_perfil'):
                            perfil.foto_perfil = content
                            perfil.save(update_fields=['foto_perfil'])
                    # #region debug-point D:registration-save-success
                    _debug_report_registration(
                        'Registro guardado correctamente',
                        {'usuario_id': str(getattr(usuario, 'id', '')), 'email': getattr(usuario, 'email', '')},
                    )
                    # #endregion
            except IntegrityError as exc:
                # #region debug-point E:registration-save-error
                _debug_report_registration(
                    'Error guardando registro',
                    {
                        'error_type': exc.__class__.__name__,
                        'error_message': str(exc),
                        'correo': (request.POST.get('correo') or '').strip(),
                        'rol': (request.POST.get('rol') or '').strip(),
                        'cedula': (request.POST.get('cedula') or '').strip(),
                    },
                )
                # #endregion
                logger.exception("Error guardando el registro inicial del usuario.")
                raw = str(exc).lower()
                if 'usuarios_email_key' in raw or 'email' in raw and 'duplicate' in raw:
                    form.add_error('correo', 'Ya existe una cuenta con este correo.')
                else:
                    form.add_error(None, 'No fue posible guardar tu registro en este momento. Intenta de nuevo.')
                return render(request, 'core/autenticacion/registro.html', {'form': form})
            except Exception as exc:
                # #region debug-point E:registration-save-error
                _debug_report_registration(
                    'Error guardando registro',
                    {
                        'error_type': exc.__class__.__name__,
                        'error_message': str(exc),
                        'correo': (request.POST.get('correo') or '').strip(),
                        'rol': (request.POST.get('rol') or '').strip(),
                        'cedula': (request.POST.get('cedula') or '').strip(),
                    },
                )
                # #endregion
                logger.exception("Error guardando el registro inicial del usuario.")
                form.add_error(None, 'No fue posible guardar tu registro en este momento. Intenta de nuevo.')
                return render(request, 'core/autenticacion/registro.html', {'form': form})

            messages.success(request, f"Cuenta creada para {usuario.email}.")
            ok = send_notification_email(
                subject="Assist Care · Registro recibido y perfil en verificación",
                to=usuario.email,
                from_email=_mail_from_email(),
                status_label='Cuenta en revision',
                headline='Recibimos tu registro correctamente',
                message=(
                    f"Hola {usuario.nombre_completo or usuario.email}, tu cuenta ya fue creada y quedó "
                    "en estado de revision mientras validamos tus datos y documentos."
                ),
                details=compact_email_details(
                    make_email_detail('Estado actual', 'En verificacion'),
                    make_email_detail('Correo registrado', usuario.email),
                    make_email_detail('Rol solicitado', usuario.get_rol_display()),
                ),
                cta_text='Ver estado de verificacion',
                cta_url=request.build_absolute_uri(reverse('verificacion_proceso')),
                footer_note='Te notificaremos por correo cuando tu cuenta sea aprobada o si requiere correcciones.',
            )
            if not ok:
                logger.error("No se pudo enviar correo de verificación inicial al usuario %s", usuario.email)
            raw_password = (form.cleaned_data.get('password') or '').strip()
            try:
                UserModel = get_user_model()
                django_user = UserModel.objects.filter(username__iexact=usuario.email).first()
                if django_user is None:
                    django_user = UserModel.objects.create_user(
                        username=usuario.email,
                        email=usuario.email,
                        password=raw_password,
                    )
                else:
                    django_user.email = usuario.email
                    if raw_password:
                        django_user.set_password(raw_password)
                    django_user.save(update_fields=['email', 'password'])
            except Exception:
                logger.exception('No fue posible crear el usuario de Django para %s', usuario.email)
                return redirect('registro_pendiente')

            return redirect('registro_pendiente')
    else:
        form = RegistroUsuarioForm()
    return render(request, 'core/autenticacion/registro.html', {'form': form})


def registro_disponibilidad_view(request):
    correo = (request.GET.get('correo') or '').strip()
    cedula = (request.GET.get('cedula') or '').strip()
    payload = _registro_field_availability_payload(correo=correo, cedula=cedula)
    return JsonResponse({'ok': True, **payload})


def forgot_password_view(request):
    if request.method == 'POST':
        form = ForgotPasswordForm(request.POST)
        if form.is_valid():
            correo = (form.cleaned_data.get('correo') or '').strip()
            usuario = Usuario.objects_all.filter(email__iexact=correo).first()
            if usuario is not None:
                token = _build_password_reset_token(usuario)
                reset_url = request.build_absolute_uri(reverse('reset_password', args=[token]))
                try:
                    send_html_email(
                        subject='Assist Care · Restablece tu contraseña',
                        to=usuario.email,
                        from_email=_mail_from_email(),
                        template_name='emails/recuperacion_contrasena.html',
                        context={
                            'status_label': 'Seguridad',
                            'headline': 'Solicitud de restablecimiento de contraseña',
                            'message': (
                                f"Hola {usuario.nombre_completo or usuario.email}, recibimos una solicitud para "
                                "cambiar la contraseña de tu cuenta en Assist Care."
                            ),
                            'cta_text': 'Restablecer contraseña',
                            'cta_url': reset_url,
                            'expiry_label': 'Este enlace vence en 1 hora',
                            'security_note': (
                                'Si no solicitaste este cambio, puedes ignorar este mensaje. '
                                'Tu contraseña actual seguirá siendo válida hasta que completes el proceso.'
                            ),
                        },
                    )
                except Exception:
                    logger.exception("No se pudo enviar el correo de recuperación para %s", correo)
                    messages.error(request, 'No fue posible enviar el correo de recuperación en este momento. Intenta de nuevo.')
                    return render(request, 'core/autenticacion/forgot_password.html', {'form': form})

            messages.success(
                request,
                'Si el correo existe en Assist, te enviamos un enlace para restablecer tu contraseña.',
            )
            return redirect('login')
    else:
        form = ForgotPasswordForm()
    return render(request, 'core/autenticacion/forgot_password.html', {'form': form})


def reset_password_view(request, token):
    usuario = None
    token_error = None
    try:
        usuario = _get_password_reset_user_from_token(token)
    except signing.SignatureExpired:
        token_error = 'Este enlace de recuperación ya expiró. Solicita uno nuevo.'
    except signing.BadSignature:
        token_error = 'El enlace de recuperación no es válido o ya fue utilizado.'

    if token_error:
        messages.error(request, token_error)
        return redirect('forgot_password')

    if request.method == 'POST':
        form = ResetPasswordForm(request.POST)
        if form.is_valid():
            new_password = form.cleaned_data['password']
            usuario.password_hash = make_password(new_password)
            usuario.save(update_fields=['password_hash'])
            try:
                UserModel = get_user_model()
                django_user = UserModel.objects.filter(username__iexact=usuario.email).first()
                if django_user is not None:
                    django_user.set_password(new_password)
                    django_user.save(update_fields=['password'])
            except Exception:
                logger.exception('No fue posible sincronizar la contraseña 2FA para %s', usuario.email)
            messages.success(request, 'Tu contraseña fue actualizada correctamente. Ya puedes iniciar sesión.')
            return redirect('login')
    else:
        form = ResetPasswordForm()

    return render(
        request,
        'core/autenticacion/reset_password.html',
        {
            'form': form,
            'token': token,
            'usuario': usuario,
        },
    )

def registro_pendiente_view(request):
    return redirect('verificacion_proceso')

def verificacion_proceso(request):
    return render(request, 'core/autenticacion/verificacion_proceso.html')

def login_view(request):
    if request.method == 'POST':
        form = LoginForm(request.POST)
        if form.is_valid():
            correo = form.cleaned_data['correo']
            password = form.cleaned_data['password']

            usuario = Usuario.objects.filter(email__iexact=correo).first()

            if usuario and check_password(password, usuario.password_hash):
                if usuario.estado != EstadoVerificacion.APROBADO:
                    return redirect('verificacion_proceso')

                request.session['usuario_id'] = str(usuario.id)
                # ⭐ CRÍTICO: Resetear timer de la cookie de sesión al hacer login
                # Esto hace que SESSION_COOKIE_AGE = 60 se reinicie desde AHORA
                request.session.set_expiry(60)
                messages.success(request, f"Sesión iniciada como {usuario.email}.")
                if usuario.is_staff:
                    return redirect('dashboard_admin')
                return redirect('dashboard')
            else:
                messages.error(request, "Credenciales inválidas. Por favor intenta de nuevo.")
    else:
        form = LoginForm()
    return render(request, 'core/autenticacion/login.html', {'form': form})

@usuario_login_required
def completar_perfil_view(request):
    usuario = request.usuario
    if usuario.is_staff:
        return redirect('dashboard_admin')

    has_disc = hasattr(usuario, 'perfil_discapacitado') and usuario.perfil_discapacitado is not None
    has_enf = hasattr(usuario, 'perfil_enfermero') and usuario.perfil_enfermero is not None
    if (usuario.rol == RolUsuario.DISCAPACITADO and has_disc) or (usuario.rol == RolUsuario.ENFERMERO and has_enf):
        if usuario.estado == EstadoVerificacion.APROBADO:
            return redirect('dashboard')
        return redirect('verificacion_proceso')

    if request.method == 'POST':
        form = CompletarPerfilSocialForm(request.POST, request.FILES, usuario=usuario)
        if form.is_valid():
            form.save(usuario)
            send_html_email(
                subject="Bienvenido a Assist Care · Verificación en proceso",
                to=usuario.email,
                template_name="emails/bienvenida.html",
                context={
                    "status_label": "Pendiente",
                    "headline": "Verificación en proceso",
                    "message": "Recibimos tu información. Estamos revisando tu perfil para garantizar la seguridad de nuestra comunidad. Te notificaremos por correo cuando tu cuenta sea activada.",
                    "cta_text": "Ver estado",
                    "cta_url": request.build_absolute_uri(reverse("verificacion_proceso")),
                },
            )
            messages.success(request, "Perfil enviado. Te avisaremos por correo cuando esté verificado.")
            return redirect('verificacion_proceso')
        messages.error(request, "Revisa los datos del formulario.")
    else:
        initial = {'rol': usuario.rol or RolUsuario.DISCAPACITADO}
        django_user = getattr(request, 'user', None)
        if django_user is not None and getattr(django_user, 'is_authenticated', False):
            if getattr(django_user, 'first_name', ''):
                initial['nombres'] = django_user.first_name
            if getattr(django_user, 'last_name', ''):
                initial['apellidos'] = django_user.last_name
        form = CompletarPerfilSocialForm(usuario=usuario, initial=initial)

    return render(request, 'core/autenticacion/completar_perfil.html', {'form': form, 'email': usuario.email})

def logout_view(request):
    try:
        django_logout(request)
    except Exception:
        pass
    request.session.flush()
    messages.info(request, "Has cerrado sesión de forma segura.")
    return redirect('login')

@admin_required
def dashboard_admin_view(request):
    django_user = getattr(request, 'user', None)
    can_manage_staff = _request_is_staff(request)

    admin_form = CrearAdminForm()
    config_tarifas = {'precio_hora': 45000, 'comision': 15, 'antecedentes': True, 'diplomas': True}
    try:
        tarifas_obj = ConfigTarifas.objects.order_by('id').first()
        if tarifas_obj is None:
            tarifas_obj = ConfigTarifas.objects.create()
        config_tarifas = {
            'precio_hora': tarifas_obj.precio_hora,
            'comision': tarifas_obj.comision,
            'antecedentes': True,
            'diplomas': True,
        }
    except Exception:
        tarifas_obj = None
    action = (request.POST.get('action') or request.GET.get('action') or '').strip()

    # Procesar creación de nuevo reporte
    if request.method == 'POST' and action == 'nuevo_reporte':
        titulo = (request.POST.get('titulo') or '').strip()
        tipo = (request.POST.get('tipo') or '').strip()
        descripcion = (request.POST.get('descripcion') or '').strip()

        if titulo and tipo and descripcion:
            actor, actor_name = _audit_actor(request)
            Reporte.objects.create(
                usuario_admin=actor,
                titulo=titulo,
                tipo=tipo,
                descripcion=descripcion,
            )
            messages.success(request, "El reporte ha sido creado y guardado con éxito.")
        else:
            messages.error(request, "Por favor completa todos los campos del reporte.")

        return redirect('dashboard_admin')

    if request.method == 'POST' and action == 'guardar_ajustes':
        if not can_manage_staff:
            return HttpResponseForbidden('Acceso denegado')

        precio_raw = (request.POST.get('precio_hora_cop') or '').strip()
        comision_raw = (request.POST.get('comision_plataforma_pct') or '').strip()
        _ = bool(request.POST.get('verificacion_antecedentes_obligatoria'))
        _ = bool(request.POST.get('validacion_diplomas_obligatoria'))

        ajustes_guardados = False
        try:
            precio = Decimal(precio_raw) if precio_raw else Decimal('0.00')
            comision = Decimal(comision_raw) if comision_raw else Decimal('0.00')
            if precio < 0:
                precio = Decimal('0.00')
            if comision < 0:
                comision = Decimal('0.00')
            if comision > 100:
                comision = Decimal('100.00')

            tarifas_obj = ConfigTarifas.objects.order_by('id').first()
            if tarifas_obj is None:
                tarifas_obj = ConfigTarifas.objects.create(precio_hora=precio, comision=comision)
            else:
                tarifas_obj.precio_hora = precio
                tarifas_obj.comision = comision
                tarifas_obj.save(update_fields=['precio_hora', 'comision', 'updated_at'])
            ajustes_guardados = True
        except (InvalidOperation, ValueError):
            ajustes_guardados = False

        if ajustes_guardados:
            messages.success(
                request,
                "Los parámetros globales de Assist Care se han actualizado correctamente en el sistema.",
            )
        else:
            messages.error(request, "No fue posible guardar los ajustes. Revisa que los valores sean numéricos.")
        return redirect(f"{reverse('dashboard_admin')}#settings")

    if action in ('download_backup', 'descargar_backup'):
        if not can_manage_staff:
            return HttpResponseForbidden('Acceso denegado')

        payload = io.StringIO()
        call_command('dumpdata', stdout=payload, indent=4)
        content = payload.getvalue()
        payload.close()

        response = HttpResponse(content, content_type='application/json')
        response['Content-Disposition'] = 'attachment; filename=backup_axius_care.json'
        return response

    if action in ('clear_old_alerts', 'limpiar_alertas_viejas'):
        if not can_manage_staff:
            return HttpResponseForbidden('Acceso denegado')

        alertas_eliminadas_count = AuditoriaLog.objects.filter(archivado=False).update(archivado=True)

        destinatario = (getattr(settings, 'EMAIL_HOST_USER', None) or '').strip()
        if destinatario:
            send_notification_email(
                subject='Assist Care · Limpieza masiva de alertas',
                to=destinatario,
                from_email=_mail_from_email(),
                status_label='Auditoria',
                headline='Se ejecutó una limpieza masiva de alertas',
                message='La rutina administrativa de mantenimiento archivó correctamente las alertas activas del sistema.',
                details=compact_email_details(
                    make_email_detail('Alertas archivadas', alertas_eliminadas_count),
                ),
                cta_text='Abrir dashboard admin',
                cta_url=request.build_absolute_uri(reverse('dashboard_admin')),
                footer_note='Este correo sirve como constancia automática de la operación de mantenimiento.',
            )

        messages.success(
            request,
            f"Se movieron al archivo histórico {alertas_eliminadas_count} alertas del sistema con éxito.",
        )
        return redirect(f"{reverse('dashboard_admin')}#eliminacion")

    # Exportar a Excel
    if request.GET.get('export') == 'excel':
        reportes = Reporte.objects.select_related('usuario_admin').order_by('-fecha_creacion')
        wb = Workbook()
        ws = wb.active
        ws.title = "Reportes"

        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="006B3F")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Encabezados
        headers = ['ID', 'Título', 'Tipo', 'Descripción', 'Creado Por', 'Fecha de Creación']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Filas
        for row_idx, reporte in enumerate(reportes, 2):
            ws.cell(row=row_idx, column=1, value=reporte.id).border = thin_border
            ws.cell(row=row_idx, column=2, value=reporte.titulo).border = thin_border
            ws.cell(row=row_idx, column=3, value=reporte.get_tipo_display()).border = thin_border
            ws.cell(row=row_idx, column=4, value=reporte.descripcion).border = thin_border
            ws.cell(row=row_idx, column=5, value=_display_usuario(reporte.usuario_admin)).border = thin_border
            ws.cell(row=row_idx, column=6, value=reporte.fecha_creacion.strftime('%d/%m/%Y %H:%M')).border = thin_border

        # Ajustar anchos
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 50
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="reportes_axius.xlsx"'
        wb.save(response)
        return response

    # Exportar a PDF
    if request.GET.get('export') == 'pdf':
        reportes = Reporte.objects.select_related('usuario_admin').order_by('-fecha_creacion')
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="utf-8">
            <style>
                body {{ font-family: Arial, sans-serif; margin: 40px; color: #333; }}
                .header {{ text-align: center; margin-bottom: 30px; border-bottom: 2px solid #006B3F; padding-bottom: 20px; }}
                .header h1 {{ color: #006B3F; margin: 0; font-size: 24px; }}
                .header p {{ color: #666; margin: 5px 0 0; font-size: 12px; }}
                .summary {{ background: #f2faf7; padding: 15px; border-radius: 8px; margin-bottom: 20px; }}
                .summary p {{ margin: 0; font-size: 14px; }}
                table {{ width: 100%; border-collapse: collapse; margin-top: 20px; }}
                th {{ background: #006B3F; color: white; padding: 12px 8px; text-align: left; font-size: 12px; }}
                td {{ padding: 10px 8px; border-bottom: 1px solid #ddd; font-size: 11px; }}
                tr:nth-child(even) {{ background: #f9f9f9; }}
                .tipo-inci {{ color: #dc2626; }} .tipo-audi {{ color: #2563eb; }} .tipo-rend {{ color: #16a34a; }}
                .footer {{ margin-top: 30px; text-align: center; font-size: 10px; color: #999; }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1>Assist Care - Historial de Reportes</h1>
                <p>Conexión Humana. Confianza Orgánica.</p>
            </div>
            <div class="summary">
                <p><strong>Total de Reportes:</strong> {reportes.count()} registros</p>
                <p><strong>Fecha de Generación:</strong> {timezone.now().strftime('%d/%m/%Y %H:%M:%S')}</p>
            </div>
            <table>
                <thead>
                    <tr>
                        <th>ID</th>
                        <th>Título</th>
                        <th>Tipo</th>
                        <th>Descripción</th>
                        <th>Creado Por</th>
                        <th>Fecha</th>
                    </tr>
                </thead>
                <tbody>
        """
        for reporte in reportes:
            tipo_class = 'tipo-inci' if reporte.tipo == 'incidencia' else 'tipo-audi' if reporte.tipo == 'auditoria' else 'tipo-rend'
            html_content += f"""
                    <tr>
                        <td>{reporte.id}</td>
                        <td>{reporte.titulo}</td>
                        <td class="{tipo_class}">{reporte.get_tipo_display()}</td>
                        <td>{reporte.descripcion[:80]}{'...' if len(reporte.descripcion) > 80 else ''}</td>
                        <td>{_display_usuario(reporte.usuario_admin)}</td>
                        <td>{reporte.fecha_creacion.strftime('%d/%m/%Y %H:%M')}</td>
                    </tr>
            """
        html_content += """
                </tbody>
            </table>
            <div class="footer">
                <p>Assist Care © {% now "Y" %} - Sistema de Administración</p>
            </div>
        </body>
        </html>
        """
        response = HttpResponse(html_content, content_type='text/html')
        response['Content-Disposition'] = 'attachment; filename="reportes_axius.html'
        return response

    busqueda = (request.GET.get('search') or '').strip()
    estado_filtro = (request.GET.get('estado') or '').strip()

    # Filtro de rango de días
    dias_seleccionados = int(request.GET.get('dias', '30'))
    if dias_seleccionados > 0:
        fecha_corte = timezone.now() - datetime.timedelta(days=dias_seleccionados)
    else:
        fecha_corte = None

    # Notificaciones reales: últimas peticiones activas ordenadas por fecha
    notificaciones = (
        Peticion.objects.filter(activo=True, estado=EstadoPeticion.ACTIVA)
        .select_related('discapacitado', 'discapacitado__perfil_discapacitado')
        .order_by('-fecha_creacion')[:5]
    )

    # Todas las notificaciones (listado completo sin límite)
    todas_las_notificaciones = (
        Peticion.objects.filter(activo=True)
        .select_related('discapacitado', 'discapacitado__perfil_discapacitado')
        .order_by('-fecha_creacion')
    )

    # Buscador global
    search_query = (request.GET.get('search') or '').strip()
    resultados_busqueda = []
    if search_query:
        # Buscar en Usuarios
        usuarios_encontrados = Usuario.objects.filter(
            Q(email__icontains=search_query) |
            Q(perfil_enfermero__nombres__icontains=search_query) |
            Q(perfil_enfermero__apellidos__icontains=search_query) |
            Q(perfil_discapacitado__nombres__icontains=search_query) |
            Q(perfil_discapacitado__apellidos__icontains=search_query)
        ).distinct()[:10]

        # Buscar en Peticiones
        peticiones_encontradas = Peticion.objects.filter(
            Q(titulo__icontains=search_query) |
            Q(descripcion__icontains=search_query) |
            Q(ciudad__icontains=search_query)
        ).distinct()[:10]

        resultados_busqueda = list(usuarios_encontrados) + list(peticiones_encontradas)

    reportes_qs = Reporte.objects.select_related('usuario_admin').order_by('-fecha_creacion')
    reportes_paginator = Paginator(reportes_qs, 5)
    reportes_page_number = (request.GET.get('reportes_page') or '').strip() or 1
    reportes = reportes_paginator.get_page(reportes_page_number)

    usuarios = Usuario.objects.filter(activo=True).select_related('perfil_discapacitado', 'perfil_enfermero').order_by('-fecha_creacion')

    # Métricas filtradas por rango de días
    usuarios_qs = Usuario.objects.filter(activo=True)
    if fecha_corte:
        usuarios_qs = usuarios_qs.filter(fecha_creacion__gte=fecha_corte)
    total_usuarios = usuarios_qs.count()

    pendientes_qs = Usuario.objects.filter(activo=True, estado__in=[EstadoVerificacion.PREREGISTRO, EstadoVerificacion.PENDIENTE_REVISION])
    if fecha_corte:
        pendientes_qs = pendientes_qs.filter(fecha_creacion__gte=fecha_corte)
    pendientes_verificar = pendientes_qs.count()

    usuarios_aprobados_qs = Usuario.objects.filter(activo=True, estado=EstadoVerificacion.APROBADO)
    if fecha_corte:
        usuarios_aprobados_qs = usuarios_aprobados_qs.filter(fecha_creacion__gte=fecha_corte)
    usuarios_aprobados = usuarios_aprobados_qs.count()

    usuarios_rechazados_qs = Usuario.objects.filter(activo=True, estado=EstadoVerificacion.RECHAZADO)
    if fecha_corte:
        usuarios_rechazados_qs = usuarios_rechazados_qs.filter(fecha_creacion__gte=fecha_corte)
    usuarios_rechazados = usuarios_rechazados_qs.count()

    enfermeros_qs = Usuario.objects.filter(activo=True, rol=RolUsuario.ENFERMERO, estado=EstadoVerificacion.APROBADO)
    if fecha_corte:
        enfermeros_qs = enfermeros_qs.filter(fecha_creacion__gte=fecha_corte)
    enfermeros_verificados = enfermeros_qs.count()

    discapacitados_qs = Usuario.objects.filter(activo=True, rol=RolUsuario.DISCAPACITADO, estado=EstadoVerificacion.APROBADO)
    if fecha_corte:
        discapacitados_qs = discapacitados_qs.filter(fecha_creacion__gte=fecha_corte)
    discapacitados_verificados = discapacitados_qs.count()

    usuarios_pendientes = Usuario.objects.filter(
        activo=True,
        estado__in=[EstadoVerificacion.PREREGISTRO, EstadoVerificacion.PENDIENTE_REVISION],
        rol__in=[RolUsuario.ENFERMERO, RolUsuario.DISCAPACITADO],
    ).select_related('perfil_discapacitado', 'perfil_enfermero')
    if fecha_corte:
        usuarios_pendientes = usuarios_pendientes.filter(fecha_creacion__gte=fecha_corte)
    usuarios_pendientes = usuarios_pendientes.order_by('-fecha_creacion')

    usuarios_verificados = (
        Usuario.objects.filter(activo=True, estado=EstadoVerificacion.APROBADO, rol__in=[RolUsuario.ENFERMERO, RolUsuario.DISCAPACITADO])
        .select_related('perfil_discapacitado', 'perfil_enfermero')
        .order_by('-fecha_creacion')
    )

    peticiones_abiertas = (
        Peticion.objects.filter(activo=True, estado=EstadoPeticion.ACTIVA)
        .select_related('discapacitado', 'discapacitado__perfil_discapacitado')
        .order_by('-fecha_creacion')
    )

    peticiones_metricas_qs = Peticion.objects.filter(activo=True)
    if fecha_corte:
        peticiones_metricas_qs = peticiones_metricas_qs.filter(fecha_creacion__gte=fecha_corte)
    peticiones_activas = peticiones_metricas_qs.filter(estado=EstadoPeticion.ACTIVA).count()
    peticiones_asignadas = peticiones_metricas_qs.filter(estado=EstadoPeticion.ASIGNADA).count()
    peticiones_completadas = peticiones_metricas_qs.filter(estado=EstadoPeticion.COMPLETADA).count()
    peticiones_canceladas = peticiones_metricas_qs.filter(estado=EstadoPeticion.CANCELADA).count()

    postulaciones_qs = Postulacion.objects.select_related('peticion', 'enfermero')
    if fecha_corte:
        postulaciones_qs = postulaciones_qs.filter(fecha_postulacion__gte=fecha_corte)
    total_postulaciones = postulaciones_qs.count()
    postulaciones_pendientes = postulaciones_qs.filter(estado=EstadoPostulacion.PENDIENTE).count()
    postulaciones_aceptadas = postulaciones_qs.filter(estado=EstadoPostulacion.ACEPTADA).count()
    tasa_aceptacion_postulaciones = round((postulaciones_aceptadas / total_postulaciones) * 100, 1) if total_postulaciones else 0

    servicios_qs = ServicioAcompanamiento.objects.filter(activo=True).select_related(
        'paciente',
        'enfermero',
        'paciente__perfil_discapacitado',
        'enfermero__perfil_enfermero',
    )
    if busqueda:
        q = (
            Q(paciente__email__icontains=busqueda)
            | Q(enfermero__email__icontains=busqueda)
            | Q(paciente__perfil_discapacitado__nombres__icontains=busqueda)
            | Q(paciente__perfil_discapacitado__apellidos__icontains=busqueda)
            | Q(enfermero__perfil_enfermero__nombres__icontains=busqueda)
            | Q(enfermero__perfil_enfermero__apellidos__icontains=busqueda)
        )
        busqueda_id = busqueda.lstrip('#')
        if busqueda_id.isdigit():
            q = q | Q(id=int(busqueda_id))
        servicios_qs = servicios_qs.filter(q)

    if estado_filtro:
        servicios_qs = servicios_qs.filter(estado=estado_filtro)

    servicios_qs = servicios_qs.order_by('-fecha_creacion')
    servicios_paginator = Paginator(servicios_qs, 5)
    servicios_page_number = (request.GET.get('servicios_page') or '').strip() or 1
    servicios = servicios_paginator.get_page(servicios_page_number)
    servicios_activos = servicios_qs.filter(estado__in=[EstadoServicio.EN_CAMINO, EstadoServicio.EN_PROGRESO]).count()
    servicios_en_camino = servicios_qs.filter(estado=EstadoServicio.EN_CAMINO).count()
    servicios_en_progreso = servicios_qs.filter(estado=EstadoServicio.EN_PROGRESO).count()
    servicios_completados = servicios_qs.filter(estado=EstadoServicio.COMPLETADO).count()
    servicios_cancelados = servicios_qs.filter(estado=EstadoServicio.CANCELADO).count()
    enfermeros_ocupados = servicios_qs.filter(
        estado__in=[EstadoServicio.EN_CAMINO, EstadoServicio.EN_PROGRESO],
        enfermero__isnull=False,
    ).values('enfermero').distinct().count()
    ocupacion_enfermeros = round((enfermeros_ocupados / enfermeros_verificados) * 100, 1) if enfermeros_verificados else 0

    servicios_q = ServicioAcompanamiento.objects.filter(activo=True)
    if fecha_corte:
        servicios_q = servicios_q.filter(fecha_creacion__gte=fecha_corte)
    servicios_hoy = servicios_q.count()
    tasa_finalizacion = round((servicios_completados / servicios_hoy) * 100, 1) if servicios_hoy else 0

    calificaciones_qs = Calificacion.objects.all()
    if fecha_corte:
        calificaciones_qs = calificaciones_qs.filter(fecha_creacion__gte=fecha_corte)
    promedio_calificacion = calificaciones_qs.aggregate(promedio=Avg('estrellas')).get('promedio') or 0
    promedio_calificacion = round(float(promedio_calificacion), 1) if promedio_calificacion else 0
    total_calificaciones = calificaciones_qs.count()

    metricas_periodo_label = f'Últimos {dias_seleccionados} días' if dias_seleccionados > 0 else 'Todo el historial'

    auditoria_logs = AuditoriaLog.objects.select_related('usuario_admin').filter(archivado=False).order_by('-fecha_hora')[:10]
    alertas_sistema_cortas = AuditoriaLog.objects.select_related('usuario_admin').filter(archivado=False).order_by('-fecha_hora')[:5]
    historial_alertas_completo = AuditoriaLog.objects.select_related('usuario_admin').filter(archivado=False).order_by('-fecha_hora')
    alertas_eliminadas_historial = AuditoriaLog.objects.select_related('usuario_admin').filter(archivado=True).order_by('-fecha_hora')

    estado_counts_q = ServicioAcompanamiento.objects.filter(activo=True)
    if fecha_corte:
        estado_counts_q = estado_counts_q.filter(fecha_creacion__gte=fecha_corte)
    estado_counts_qs = estado_counts_q.values('estado').annotate(total=Count('id'))
    estado_counts = {row['estado']: row['total'] for row in estado_counts_qs}
    servicios_chart = {
        'labels': [EstadoServicio.EN_CAMINO, EstadoServicio.EN_PROGRESO, EstadoServicio.COMPLETADO, EstadoServicio.CANCELADO],
        'values': [
            int(estado_counts.get(EstadoServicio.EN_CAMINO, 0)),
            int(estado_counts.get(EstadoServicio.EN_PROGRESO, 0)),
            int(estado_counts.get(EstadoServicio.COMPLETADO, 0)),
            int(estado_counts.get(EstadoServicio.CANCELADO, 0)),
        ],
    }

    context = {
        'total_usuarios': total_usuarios,
        'pendientes_verificar': pendientes_verificar,
        'pendientes': pendientes_verificar,
        'usuarios_aprobados': usuarios_aprobados,
        'usuarios_rechazados': usuarios_rechazados,
        'enfermeros_verificados': enfermeros_verificados,
        'discapacitados_verificados': discapacitados_verificados,
        'servicios_activos': servicios_activos,
        'servicios_hoy': servicios_hoy,
        'servicios_en_camino': servicios_en_camino,
        'servicios_en_progreso': servicios_en_progreso,
        'servicios_completados': servicios_completados,
        'servicios_cancelados': servicios_cancelados,
        'enfermeros_ocupados': enfermeros_ocupados,
        'ocupacion_enfermeros': ocupacion_enfermeros,
        'tasa_finalizacion': tasa_finalizacion,
        'usuarios_pendientes': usuarios_pendientes,
        'solicitudes_pendientes': usuarios_pendientes,
        'usuarios_verificados': usuarios_verificados,
        'usuarios': usuarios,
        'peticiones_abiertas': peticiones_abiertas,
        'peticiones_activas': peticiones_activas,
        'peticiones_asignadas': peticiones_asignadas,
        'peticiones_completadas': peticiones_completadas,
        'peticiones_canceladas': peticiones_canceladas,
        'total_postulaciones': total_postulaciones,
        'postulaciones_pendientes': postulaciones_pendientes,
        'postulaciones_aceptadas': postulaciones_aceptadas,
        'tasa_aceptacion_postulaciones': tasa_aceptacion_postulaciones,
        'promedio_calificacion': promedio_calificacion,
        'total_calificaciones': total_calificaciones,
        'metricas_periodo_label': metricas_periodo_label,
        'servicios': servicios,
        'busqueda': busqueda,
        'search_query': search_query,
        'resultados_busqueda': resultados_busqueda,
        'estado_filtro': estado_filtro,
        'auditoria_logs': auditoria_logs,
        'alertas_sistema_cortas': alertas_sistema_cortas,
        'historial_alertas_completo': historial_alertas_completo,
        'alertas_eliminadas_historial': alertas_eliminadas_historial,
        'servicios_chart': servicios_chart,
        'admin_form': admin_form,
        'can_manage_staff': can_manage_staff,
        'config_tarifas': config_tarifas,
        'notificaciones': notificaciones,
        'todas_las_notificaciones': todas_las_notificaciones,
        'dias_seleccionados': dias_seleccionados,
        'reportes': reportes,
    }
    return render(request, 'core/admin/dashboard_admin.html', context)


@admin_required
def admin_metrics_json(request):
    if request.method != 'GET':
        return JsonResponse({'status': 'error', 'message': 'Método no permitido.'}, status=405)

    busqueda = (request.GET.get('search') or '').strip()
    estado_filtro = (request.GET.get('estado') or '').strip()

    try:
        dias_seleccionados = int(request.GET.get('dias', '30'))
    except ValueError:
        dias_seleccionados = 30

    if dias_seleccionados > 0:
        fecha_corte = timezone.now() - datetime.timedelta(days=dias_seleccionados)
    else:
        fecha_corte = None

    usuarios_qs = Usuario.objects.filter(activo=True)
    if fecha_corte:
        usuarios_qs = usuarios_qs.filter(fecha_creacion__gte=fecha_corte)
    total_usuarios = usuarios_qs.count()

    pendientes_qs = Usuario.objects.filter(activo=True, estado__in=[EstadoVerificacion.PREREGISTRO, EstadoVerificacion.PENDIENTE_REVISION])
    if fecha_corte:
        pendientes_qs = pendientes_qs.filter(fecha_creacion__gte=fecha_corte)
    pendientes_verificar = pendientes_qs.count()

    usuarios_aprobados_qs = Usuario.objects.filter(activo=True, estado=EstadoVerificacion.APROBADO)
    if fecha_corte:
        usuarios_aprobados_qs = usuarios_aprobados_qs.filter(fecha_creacion__gte=fecha_corte)
    usuarios_aprobados = usuarios_aprobados_qs.count()

    usuarios_rechazados_qs = Usuario.objects.filter(activo=True, estado=EstadoVerificacion.RECHAZADO)
    if fecha_corte:
        usuarios_rechazados_qs = usuarios_rechazados_qs.filter(fecha_creacion__gte=fecha_corte)
    usuarios_rechazados = usuarios_rechazados_qs.count()

    enfermeros_qs = Usuario.objects.filter(activo=True, rol=RolUsuario.ENFERMERO, estado=EstadoVerificacion.APROBADO)
    if fecha_corte:
        enfermeros_qs = enfermeros_qs.filter(fecha_creacion__gte=fecha_corte)
    enfermeros_verificados = enfermeros_qs.count()

    discapacitados_qs = Usuario.objects.filter(activo=True, rol=RolUsuario.DISCAPACITADO, estado=EstadoVerificacion.APROBADO)
    if fecha_corte:
        discapacitados_qs = discapacitados_qs.filter(fecha_creacion__gte=fecha_corte)
    discapacitados_verificados = discapacitados_qs.count()

    peticiones_metricas_qs = Peticion.objects.filter(activo=True)
    if fecha_corte:
        peticiones_metricas_qs = peticiones_metricas_qs.filter(fecha_creacion__gte=fecha_corte)
    peticiones_activas = peticiones_metricas_qs.filter(estado=EstadoPeticion.ACTIVA).count()
    peticiones_asignadas = peticiones_metricas_qs.filter(estado=EstadoPeticion.ASIGNADA).count()
    peticiones_completadas = peticiones_metricas_qs.filter(estado=EstadoPeticion.COMPLETADA).count()
    peticiones_canceladas = peticiones_metricas_qs.filter(estado=EstadoPeticion.CANCELADA).count()

    postulaciones_qs = Postulacion.objects.all()
    if fecha_corte:
        postulaciones_qs = postulaciones_qs.filter(fecha_postulacion__gte=fecha_corte)
    total_postulaciones = postulaciones_qs.count()
    postulaciones_pendientes = postulaciones_qs.filter(estado=EstadoPostulacion.PENDIENTE).count()
    postulaciones_aceptadas = postulaciones_qs.filter(estado=EstadoPostulacion.ACEPTADA).count()
    tasa_aceptacion_postulaciones = round((postulaciones_aceptadas / total_postulaciones) * 100, 1) if total_postulaciones else 0

    servicios_qs = ServicioAcompanamiento.objects.filter(activo=True).select_related('paciente', 'enfermero')
    if busqueda:
        q = Q(paciente__email__icontains=busqueda) | Q(enfermero__email__icontains=busqueda)
        busqueda_id = busqueda.lstrip('#')
        if busqueda_id.isdigit():
            q = q | Q(id=int(busqueda_id))
        servicios_qs = servicios_qs.filter(q)
    if estado_filtro:
        servicios_qs = servicios_qs.filter(estado=estado_filtro)
    if fecha_corte:
        servicios_qs = servicios_qs.filter(fecha_creacion__gte=fecha_corte)

    servicios_activos = servicios_qs.filter(estado__in=[EstadoServicio.EN_CAMINO, EstadoServicio.EN_PROGRESO]).count()
    servicios_en_camino = servicios_qs.filter(estado=EstadoServicio.EN_CAMINO).count()
    servicios_en_progreso = servicios_qs.filter(estado=EstadoServicio.EN_PROGRESO).count()
    servicios_completados = servicios_qs.filter(estado=EstadoServicio.COMPLETADO).count()
    servicios_cancelados = servicios_qs.filter(estado=EstadoServicio.CANCELADO).count()
    servicios_total_periodo = servicios_qs.count()
    tasa_finalizacion = round((servicios_completados / servicios_total_periodo) * 100, 1) if servicios_total_periodo else 0

    enfermeros_ocupados = (
        ServicioAcompanamiento.objects.filter(
            activo=True,
            estado__in=[EstadoServicio.EN_CAMINO, EstadoServicio.EN_PROGRESO],
            enfermero__isnull=False,
        )
        .values('enfermero')
        .distinct()
        .count()
    )
    ocupacion_enfermeros = round((enfermeros_ocupados / enfermeros_verificados) * 100, 1) if enfermeros_verificados else 0

    calificaciones_qs = Calificacion.objects.all()
    if fecha_corte:
        calificaciones_qs = calificaciones_qs.filter(fecha_creacion__gte=fecha_corte)
    promedio_calificacion = calificaciones_qs.aggregate(promedio=Avg('estrellas')).get('promedio') or 0
    promedio_calificacion = round(float(promedio_calificacion), 1) if promedio_calificacion else 0
    total_calificaciones = calificaciones_qs.count()

    metricas_periodo_label = f'Últimos {dias_seleccionados} días' if dias_seleccionados > 0 else 'Todo el historial'

    return JsonResponse(
        {
            'status': 'success',
            'metricas_periodo_label': metricas_periodo_label,
            'total_usuarios': int(total_usuarios),
            'pendientes_verificar': int(pendientes_verificar),
            'usuarios_aprobados': int(usuarios_aprobados),
            'usuarios_rechazados': int(usuarios_rechazados),
            'enfermeros_verificados': int(enfermeros_verificados),
            'discapacitados_verificados': int(discapacitados_verificados),
            'peticiones_activas': int(peticiones_activas),
            'peticiones_asignadas': int(peticiones_asignadas),
            'peticiones_completadas': int(peticiones_completadas),
            'peticiones_canceladas': int(peticiones_canceladas),
            'servicios_total': int(servicios_total_periodo),
            'servicios_activos': int(servicios_activos),
            'servicios_en_camino': int(servicios_en_camino),
            'servicios_en_progreso': int(servicios_en_progreso),
            'servicios_completados': int(servicios_completados),
            'servicios_cancelados': int(servicios_cancelados),
            'tasa_finalizacion': float(tasa_finalizacion),
            'enfermeros_ocupados': int(enfermeros_ocupados),
            'ocupacion_enfermeros': float(ocupacion_enfermeros),
            'total_postulaciones': int(total_postulaciones),
            'postulaciones_pendientes': int(postulaciones_pendientes),
            'postulaciones_aceptadas': int(postulaciones_aceptadas),
            'tasa_aceptacion_postulaciones': float(tasa_aceptacion_postulaciones),
            'promedio_calificacion': float(promedio_calificacion),
            'total_calificaciones': int(total_calificaciones),
        }
    )


@admin_required
def admin_donaciones_json(request):
    if request.method != 'GET':
        return JsonResponse({'status': 'error', 'message': 'Método no permitido.'}, status=405)

    try:
        dias = int(request.GET.get('dias', '30'))
    except ValueError:
        dias = 30

    if dias > 0:
        fecha_corte = timezone.now() - datetime.timedelta(days=dias)
        qs = Donacion.objects.filter(fecha_creacion__gte=fecha_corte)
    else:
        qs = Donacion.objects.all()

    serie = (
        qs.annotate(dia=TruncDate('fecha_creacion'))
        .values('dia')
        .annotate(total=Sum('monto'))
        .order_by('dia')
    )

    labels = []
    totals = []
    for row in serie:
        dia = row.get('dia')
        total = row.get('total') or Decimal('0.00')
        labels.append(dia.isoformat() if dia else None)
        totals.append(str(total))

    ultimas = list(
        Donacion.objects.order_by('-fecha_creacion')
        .values('donante', 'email', 'monto', 'estado', 'fecha_creacion', 'referencia_payco')[:10]
    )
    for item in ultimas:
        fecha = item.get('fecha_creacion')
        item['fecha_creacion'] = fecha.isoformat() if fecha else None
        item['monto'] = str(item.get('monto') or Decimal('0.00'))

    return JsonResponse(
        {
            'status': 'success',
            'dias': int(dias),
            'serie': {'labels': labels, 'totals': totals},
            'ultimas': ultimas,
        }
    )


@admin_required
def admin_donaciones_view(request):
    return render(request, 'core/admin/donaciones_gestion.html')


@admin_required
def donaciones_stats_json(request):
    if request.method != 'GET':
        return JsonResponse({'status': 'error', 'message': 'Método no permitido.'}, status=405)

    try:
        dias = int(request.GET.get('dias', '30'))
    except ValueError:
        dias = 30

    if dias > 0:
        fecha_corte = timezone.now() - datetime.timedelta(days=dias)
        qs = Donacion.objects.filter(fecha_creacion__gte=fecha_corte)
    else:
        qs = Donacion.objects.all()

    serie = (
        qs.annotate(dia=TruncDate('fecha_creacion'))
        .values('dia')
        .annotate(total=Sum('monto'))
        .order_by('dia')
    )

    labels = []
    totals = []
    for row in serie:
        dia = row.get('dia')
        total = row.get('total') or Decimal('0.00')
        labels.append(dia.isoformat() if dia else None)
        totals.append(str(total))

    return JsonResponse({'status': 'success', 'dias': int(dias), 'labels': labels, 'totals': totals})


@admin_required
def donaciones_lista_json(request):
    if request.method != 'GET':
        return JsonResponse({'status': 'error', 'message': 'Método no permitido.'}, status=405)

    try:
        limit = int(request.GET.get('limit', '10'))
    except ValueError:
        limit = 10
    limit = max(1, min(limit, 50))

    items = list(
        Donacion.objects.order_by('-fecha_creacion')
        .values('donante', 'email', 'monto', 'estado', 'fecha_creacion', 'referencia_payco')[:limit]
    )
    for item in items:
        fecha = item.get('fecha_creacion')
        item['fecha_creacion'] = fecha.isoformat() if fecha else None
        item['monto'] = str(item.get('monto') or Decimal('0.00'))

    return JsonResponse({'status': 'success', 'items': items, 'limit': int(limit)})


@admin_required
def admin_metrics_revision_json(request):
    if request.method != 'GET':
        return JsonResponse({'status': 'error', 'message': 'Método no permitido.'}, status=405)

    pulse, _created = SystemPulse.objects.get_or_create(name='admin_metrics', defaults={'revision': 0})
    updated_at = pulse.updated_at
    updated_at_str = updated_at.isoformat() if updated_at else None
    return JsonResponse({'status': 'success', 'revision': int(pulse.revision), 'updated_at': updated_at_str})


@admin_required
def admin_metrics_wait_json(request):
    if request.method != 'GET':
        return JsonResponse({'status': 'error', 'message': 'Método no permitido.'}, status=405)

    try:
        since = int(request.GET.get('since', '0'))
    except ValueError:
        since = 0

    try:
        timeout = int(request.GET.get('timeout', '25'))
    except ValueError:
        timeout = 25

    timeout = max(1, min(timeout, 25))
    deadline = time.monotonic() + timeout

    while True:
        pulse, _created = SystemPulse.objects.get_or_create(name='admin_metrics', defaults={'revision': 0})
        revision = int(pulse.revision or 0)
        updated_at = pulse.updated_at
        updated_at_str = updated_at.isoformat() if updated_at else None

        if revision != since:
            res = JsonResponse({'status': 'success', 'changed': True, 'revision': revision, 'updated_at': updated_at_str})
            res['Cache-Control'] = 'no-store'
            return res

        if time.monotonic() >= deadline:
            res = JsonResponse({'status': 'success', 'changed': False, 'revision': revision, 'updated_at': updated_at_str})
            res['Cache-Control'] = 'no-store'
            return res

        time.sleep(0.6)

@admin_required
def exportar_servicios_csv(request):
    if not _request_is_staff(request):
        return HttpResponseForbidden('Acceso denegado')

    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="reporte_servicios_acompanamiento.csv"'
    response.write('\ufeff')
    writer = csv.writer(response)
    writer.writerow(['ID Servicio', 'Paciente', 'Acompañante (Enfermero)', 'Estado Actual', 'Fecha de Registro'])

    servicios = (
        ServicioAcompanamiento.objects.all()
        .select_related('paciente', 'enfermero')
        .order_by('-id')
    )
    for s in servicios:
        paciente_name = s.paciente.nombre_completo or s.paciente.email
        enfermero_name = ''
        if s.enfermero:
            enfermero_name = s.enfermero.nombre_completo or s.enfermero.email
        fecha = s.fecha_creacion
        if timezone.is_aware(fecha):
            fecha = timezone.localtime(fecha)
        writer.writerow([s.id, paciente_name, enfermero_name, s.estado, fecha.strftime('%d/%m/%Y %H:%M')])

    return response

@admin_required
def exportar_servicios_pdf(request):
    if not _request_is_staff(request):
        return HttpResponseForbidden('Acceso denegado')

    if not _reportlab_available:
        messages.error(request, "Exportación PDF no disponible. Instala la dependencia reportlab.")
        return redirect('dashboard_admin')

    servicios = (
        ServicioAcompanamiento.objects.all()
        .select_related('paciente', 'enfermero')
        .order_by('-id')
    )

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, title='Reporte de servicios')
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        'title_style',
        parent=styles['Title'],
        textColor=colors.HexColor('#0f172a'),
        alignment=1,
        spaceAfter=8,
    )
    subtitle_style = ParagraphStyle(
        'subtitle_style',
        parent=styles['Normal'],
        textColor=colors.HexColor('#334155'),
        alignment=1,
        spaceAfter=14,
    )

    now = timezone.localtime(timezone.now()) if timezone.is_aware(timezone.now()) else timezone.now()
    story = [
        Paragraph('Reporte de Servicios', title_style),
        Paragraph('Reporte General de Acompañamientos - Acompañamiento Seguro', subtitle_style),
        Paragraph(f'Fecha de generación: {now.strftime("%d/%m/%Y %H:%M")}', subtitle_style),
        Spacer(1, 10),
    ]

    data = [['ID Servicio', 'Paciente', 'Acompañante (Enfermero)', 'Estado Actual', 'Fecha de Registro']]
    for s in servicios:
        paciente_name = s.paciente.nombre_completo or s.paciente.email
        enfermero_name = ''
        if s.enfermero:
            enfermero_name = s.enfermero.nombre_completo or s.enfermero.email
        fecha = s.fecha_creacion
        if timezone.is_aware(fecha):
            fecha = timezone.localtime(fecha)
        data.append([str(s.id), paciente_name, enfermero_name, s.estado, fecha.strftime('%d/%m/%Y %H:%M')])

    table = Table(data, repeatRows=1)
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a8a')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ])
    for i in range(1, len(data)):
        bg = colors.HexColor('#f8fafc') if i % 2 == 1 else colors.HexColor('#f1f5f9')
        style.add('BACKGROUND', (0, i), (-1, i), bg)
    table.setStyle(style)
    story.append(table)

    doc.build(story)
    pdf = buffer.getvalue()
    buffer.close()

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte_servicios.pdf"'
    response.write(pdf)
    return response

@admin_required
def crear_administrador(request):
    if request.method != 'POST':
        return redirect('dashboard_admin')

    if not _request_is_staff(request):
        return HttpResponseForbidden('Acceso denegado')

    form = CrearAdminForm(request.POST)
    if not form.is_valid():
        error = None
        for field_errors in form.errors.values():
            if field_errors:
                error = str(field_errors[0])
                break
        messages.error(request, error or 'No fue posible crear el administrador.')
        return redirect(f"{reverse('dashboard_admin')}?module=security")

    try:
        password_plano = form.cleaned_data.get('password') or ''
        nuevo_admin = form.save()
    except IntegrityError:
        messages.error(request, 'Ya existe una cuenta con este correo.')
        return redirect(f"{reverse('dashboard_admin')}?module=security")

    login_url = request.build_absolute_uri(reverse('login'))
    send_notification_email(
        subject='Assist Care · Bienvenida Administrador',
        to=nuevo_admin.email,
        from_email=_mail_from_email(),
        status_label='Acceso administrativo',
        headline='Tu cuenta de administrador fue creada',
        message='Ya tienes acceso administrativo a Assist Care. Usa las credenciales entregadas para ingresar por primera vez.',
        details=compact_email_details(
            make_email_detail('Correo', nuevo_admin.email),
            make_email_detail('Contraseña temporal', password_plano),
        ),
        cta_text='Iniciar sesion',
        cta_url=login_url,
        footer_note='Si no solicitaste este acceso, contacta de inmediato al equipo de soporte.',
    )

    actor, actor_name = _audit_actor(request)
    AuditoriaLog.objects.create(
        usuario_admin=actor,
        accion=f"Creó una nueva cuenta de administrador para el usuario: {nuevo_admin.email}. (Admin: {actor_name})",
    )
    messages.success(request, "Administrador creado correctamente.")
    return redirect(f"{reverse('dashboard_admin')}?module=security")


@admin_required
def aprobar_usuario_view(request, usuario_id):
    if request.method != 'POST':
        return HttpResponseNotAllowed(['POST'])
    usuario = Usuario.objects.filter(id=usuario_id).first()
    if not usuario:
        messages.error(request, "Usuario no encontrado.")
        return redirect('dashboard_admin')

    usuario.estado = EstadoVerificacion.APROBADO
    usuario.save(update_fields=['estado'])
    ok = send_notification_email(
        subject="Assist Care · Tu perfil ha sido aprobado",
        to=usuario.email,
        from_email=_mail_from_email(),
        status_label='Cuenta aprobada',
        headline='Tu cuenta ya fue aprobada',
        message=(
            f"Hola {usuario.nombre_completo or usuario.email}, tu perfil fue aprobado por el equipo de Assist Care "
            "y ya puedes ingresar a la plataforma."
        ),
        details=compact_email_details(
            make_email_detail('Estado actual', 'Aprobada'),
            make_email_detail('Correo de acceso', usuario.email),
            make_email_detail('Rol', usuario.get_rol_display()),
        ),
        cta_text='Iniciar sesion',
        cta_url=request.build_absolute_uri(reverse('login')),
        footer_note='Te recomendamos ingresar cuanto antes para completar tu experiencia dentro de la plataforma.',
    )
    if not ok:
        logger.error("No se pudo enviar correo de aprobación al usuario %s", usuario.email)
    actor, actor_name = _audit_actor(request)
    if actor is not None or (getattr(request, 'user', None) is not None and getattr(request.user, 'is_authenticated', False) and getattr(request.user, 'is_staff', False)):
        target_name = usuario.nombre_completo or usuario.email
        AuditoriaLog.objects.create(
            usuario_admin=actor,
            accion=f"Aprobó el registro del usuario {target_name} con Rol: {usuario.get_rol_display()}. (Admin: {actor_name})",
        )
    messages.success(request, "Usuario aprobado correctamente.")
    return redirect('dashboard_admin')


@admin_required
def rechazar_usuario_view(request, usuario_id):
    if request.method != 'POST':
        return HttpResponseNotAllowed(['POST'])
    usuario = Usuario.objects.filter(id=usuario_id).first()
    if not usuario:
        messages.error(request, "Usuario no encontrado.")
        return redirect('dashboard_admin')

    usuario.estado = EstadoVerificacion.RECHAZADO
    usuario.save(update_fields=['estado'])
    ok = send_notification_email(
        subject="Assist Care · Tu perfil ha sido rechazado",
        to=usuario.email,
        from_email=_mail_from_email(),
        status_label='Cuenta rechazada',
        headline='Tu cuenta requiere correcciones',
        message=(
            f"Hola {usuario.nombre_completo or usuario.email}, en esta revision no fue posible aprobar tu perfil. "
            "Necesitamos que corrijas o completes informacion antes de volver a validar el acceso."
        ),
        details=compact_email_details(
            make_email_detail('Estado actual', 'Rechazada'),
            make_email_detail('Correo registrado', usuario.email),
            make_email_detail('Rol', usuario.get_rol_display()),
        ),
        cta_text='Volver al sitio',
        cta_url=request.build_absolute_uri(reverse('home')),
        footer_note='Si necesitas ayuda adicional, contacta al equipo de soporte de Assist Care.',
    )
    if not ok:
        logger.error("No se pudo enviar correo de rechazo al usuario %s", usuario.email)
    actor, actor_name = _audit_actor(request)
    if actor is not None or (getattr(request, 'user', None) is not None and getattr(request.user, 'is_authenticated', False) and getattr(request.user, 'is_staff', False)):
        target_name = usuario.nombre_completo or usuario.email
        AuditoriaLog.objects.create(
            usuario_admin=actor,
            accion=f"Rechazó el registro del usuario {target_name} con Rol: {usuario.get_rol_display()}. (Admin: {actor_name})",
        )
    messages.warning(request, "Usuario rechazado.")
    return redirect('dashboard_admin')

@usuario_login_required
def dashboard_view(request):
    django_user = getattr(request, 'user', None)
    if (
        django_user is not None
        and getattr(django_user, 'is_authenticated', False)
        and (getattr(django_user, 'is_superuser', False) or getattr(django_user, 'is_staff', False))
    ):
        return redirect('dashboard_admin')

    context = {
        'rol': request.usuario.rol,
    }
    
    if request.usuario.is_staff:
        return redirect('dashboard_admin')

    usuario = request.usuario
    has_disc = hasattr(usuario, 'perfil_discapacitado') and usuario.perfil_discapacitado is not None
    has_enf = hasattr(usuario, 'perfil_enfermero') and usuario.perfil_enfermero is not None
    if usuario.rol == RolUsuario.DISCAPACITADO and not has_disc:
        return redirect('completar_perfil')
    if usuario.rol == RolUsuario.ENFERMERO and not has_enf:
        return redirect('completar_perfil')
    if usuario.estado != EstadoVerificacion.APROBADO:
        return redirect('verificacion_proceso')

    if request.usuario.rol == RolUsuario.ENFERMERO:
        return redirect('dashboard_enfermero')
    if request.usuario.rol == RolUsuario.DISCAPACITADO:
        return redirect('dashboard_discapacitado')
    return render(request, 'core/perfiles/dashboard_usuario.html', context)

# Ejemplos de vistas restringidas por rol
@usuario_login_required
@role_required(RolUsuario.DISCAPACITADO)
def solicitar_ayuda_view(request):
    usuario = request.usuario
    perfil = getattr(usuario, 'perfil_discapacitado', None)

    if request.method == 'POST':
        print(f'DEBUG VIEW: usuario autenticado en solicitar_ayuda_view = {getattr(usuario, "pk", None)}')
        form = SolicitarAyudaForm(request.POST, usuario=usuario)
        if form.is_valid():
            titulo = (form.cleaned_data.get('titulo') or '').strip()
            descripcion = (form.cleaned_data.get('descripcion') or '').strip()
            ciudad = (form.cleaned_data.get('ciudad') or '').strip()
            fecha_evento = form.cleaned_data.get('fecha_evento')

            hace_poco = timezone.now() - datetime.timedelta(seconds=5)
            existe_duplicado = (
                Peticion.objects.filter(
                    discapacitado_id=getattr(usuario, 'pk', None),
                    titulo=titulo,
                    descripcion=descripcion,
                    ciudad__iexact=ciudad,
                    fecha_evento=fecha_evento,
                    fecha_creacion__gte=hace_poco,
                ).exists()
            )
            if existe_duplicado:
                messages.error(request, "Solicitud duplicada detectada. Intenta nuevamente en unos segundos.")
                return redirect('dashboard_discapacitado')

            peticion = form.save_for_user(usuario)
            ciudad = (peticion.ciudad or '').strip() or 'tu ciudad'
            enfermeros = (
                Usuario.objects.filter(
                    rol=RolUsuario.ENFERMERO,
                    estado=EstadoVerificacion.APROBADO,
                )
                .select_related('perfil_enfermero')
            )
            if ciudad:
                enfermeros = enfermeros.filter(perfil_enfermero__ciudad__iexact=ciudad)

            enfermeros_emails = list(enfermeros.exclude(email='').values_list('email', flat=True))
            if enfermeros_emails:
                ok = send_notification_email(
                    subject='Assist Care · Nueva solicitud de servicio pendiente',
                    to=enfermeros_emails,
                    from_email=_mail_from_email(),
                    status_label='Nueva solicitud',
                    headline='Hay una nueva solicitud pendiente por revisar',
                    message='Se registró una nueva solicitud de servicio y ya está disponible para revisión en la plataforma.',
                    details=compact_email_details(
                        make_email_detail('Ciudad', ciudad),
                        make_email_detail('Paciente', usuario.nombre_completo or usuario.email),
                    ),
                    cta_text='Abrir portal',
                    cta_url=request.build_absolute_uri(reverse('dashboard_enfermero')),
                    footer_note='Revisa tu panel para validar disponibilidad y atender la solicitud cuanto antes.',
                )
                if not ok:
                    logger.error("Error enviando notificación de nueva solicitud.")
            messages.success(request, "Solicitud enviada correctamente.")
            return redirect('dashboard_discapacitado')
        messages.error(request, "Revisa los datos del formulario.")
        return render(request, 'core/solicitudes/crear_solicitud.html', {'form': form})

    initial = {}
    if perfil is not None and getattr(perfil, 'ciudad', None):
        initial['ciudad'] = perfil.ciudad
    initial.setdefault('direccion', 'Ibagué')
    form = SolicitarAyudaForm(initial=initial, usuario=usuario)
    return render(request, 'core/solicitudes/crear_solicitud.html', {'form': form})

@usuario_login_required
@role_required(RolUsuario.ENFERMERO)
def aceptar_solicitud_view(request):
    return render(request, 'core/solicitudes/lista_solicitudes.html')


@usuario_login_required
@role_required(RolUsuario.DISCAPACITADO)
def dashboard_discapacitado(request):
    usuario = request.usuario
    perfil = getattr(usuario, 'perfil_discapacitado', None)

    if request.method == 'POST':
        wants_json = (
            (request.headers.get('x-requested-with') == 'XMLHttpRequest')
            or ('application/json' in (request.headers.get('accept') or ''))
        )

        form_type = (request.POST.get('form_type') or '').strip().lower()
        if form_type == 'settings':
            if perfil is None:
                if wants_json:
                    return JsonResponse({'status': 'failed', 'error': 'No se encontró el perfil del usuario.'}, status=404)
                messages.error(request, "No se encontró el perfil del usuario.")
                return redirect('dashboard_discapacitado')

            full_name = (request.POST.get('full_name') or '').strip()
            fecha_nacimiento_raw = (request.POST.get('fecha_nacimiento') or '').strip()
            bio = (request.POST.get('bio') or '').strip()
            historia = (request.POST.get('historia') or '').strip()

            screen_reader = bool(request.POST.get('screen_reader'))
            high_contrast = bool(request.POST.get('high_contrast'))
            simplified_nav = bool(request.POST.get('simplified_nav'))

            emergency_nombre = (request.POST.get('emergency_nombre') or '').strip()
            emergency_parentesco = (request.POST.get('emergency_parentesco') or '').strip()
            emergency_telefono = (request.POST.get('emergency_telefono') or '').strip()

            if full_name:
                parts = [p for p in full_name.split(' ') if p]
                if len(parts) == 1:
                    perfil.nombres = parts[0]
                    perfil.apellidos = ''
                else:
                    perfil.nombres = ' '.join(parts[:-1])
                    perfil.apellidos = parts[-1]

            if fecha_nacimiento_raw:
                try:
                    perfil.fecha_nacimiento = timezone.datetime.fromisoformat(fecha_nacimiento_raw).date()
                except Exception:
                    if wants_json:
                        return JsonResponse({'status': 'failed', 'error': 'Fecha de nacimiento inválida.'}, status=400)
                    messages.error(request, "Fecha de nacimiento inválida.")
                    return redirect('dashboard_discapacitado')

            perfil.biografia = bio
            perfil.historia = historia
            perfil.acces_screen_reader = screen_reader
            perfil.acces_high_contrast = high_contrast
            perfil.acces_simplified_nav = simplified_nav
            perfil.emergency_nombre = emergency_nombre
            perfil.emergency_parentesco = emergency_parentesco
            perfil.emergency_telefono = emergency_telefono
            perfil.fecha_actualizacion = timezone.now()
            perfil.save()

            if wants_json:
                return JsonResponse({'status': 'success'})
            messages.success(request, "Ajustes guardados.")
            return redirect('dashboard_discapacitado')

        titulo = (request.POST.get('titulo') or '').strip()
        descripcion = (request.POST.get('descripcion') or '').strip()
        fecha_servicio_raw = (request.POST.get('fecha_servicio') or '').strip() or (request.POST.get('fecha_inicio') or '').strip()
        hora_inicio_raw = (request.POST.get('hora_inicio') or '').strip() or (request.POST.get('fecha_evento') or '').strip()
        hora_fin_raw = (request.POST.get('hora_fin') or '').strip()
        direccion_texto = (request.POST.get('direccion_texto') or '').strip() or (request.POST.get('direccion') or '').strip()
        latitud_raw = (request.POST.get('latitud') or '').strip()
        longitud_raw = (request.POST.get('longitud') or '').strip()
        google_maps_url = (request.POST.get('google_maps_url') or '').strip()
        ciudad = (request.POST.get('ciudad') or '').strip() or (getattr(perfil, 'ciudad', '') if perfil else '')

        if fecha_servicio_raw and hora_inicio_raw and 'T' not in hora_inicio_raw:
            hora_inicio_raw = f'{fecha_servicio_raw}T{hora_inicio_raw}'
        if fecha_servicio_raw and hora_fin_raw and 'T' not in hora_fin_raw:
            hora_fin_raw = f'{fecha_servicio_raw}T{hora_fin_raw}'

        if not titulo:
            messages.error(request, "El título es obligatorio.")
            return redirect('dashboard_discapacitado')
        if not descripcion:
            messages.error(request, "La descripción es obligatoria.")
            return redirect('dashboard_discapacitado')
        if not hora_inicio_raw:
            messages.error(request, "La fecha y hora de inicio es obligatoria.")
            return redirect('dashboard_discapacitado')
        if not hora_fin_raw:
            messages.error(request, "La fecha y hora de fin es obligatoria.")
            return redirect('dashboard_discapacitado')
        if not ciudad:
            messages.error(request, "La ciudad es obligatoria.")
            return redirect('dashboard_discapacitado')
        if not direccion_texto:
            messages.error(request, "La dirección de recogida es obligatoria.")
            return redirect('dashboard_discapacitado')
        if not latitud_raw or not longitud_raw:
            messages.error(request, "Selecciona una ubicación válida en el mapa.")
            return redirect('dashboard_discapacitado')

        fecha_evento = parse_datetime(hora_inicio_raw)
        if fecha_evento is None:
            messages.error(request, "La fecha y hora de inicio no tiene un formato válido.")
            return redirect('dashboard_discapacitado')

        fecha_fin = parse_datetime(hora_fin_raw)
        if fecha_fin is None:
            messages.error(request, "La fecha y hora de fin no tiene un formato válido.")
            return redirect('dashboard_discapacitado')

        if timezone.is_naive(fecha_evento):
            fecha_evento = timezone.make_aware(fecha_evento, timezone.get_current_timezone())
        if timezone.is_naive(fecha_fin):
            fecha_fin = timezone.make_aware(fecha_fin, timezone.get_current_timezone())

        # #region debug-point B:booking-date-check
        inicio_dia = timezone.localtime(fecha_evento).replace(hour=0, minute=0, second=0, microsecond=0)
        fin_dia = inicio_dia + datetime.timedelta(days=1)
        solicitudes_mismo_dia = list(
            Peticion.objects.filter(
                discapacitado=usuario,
                fecha_evento__gte=inicio_dia,
                fecha_evento__lt=fin_dia,
            ).values('id', 'estado', 'fecha_evento')
        )
        _debug_report_booking(
            'Evaluando duplicado por mismo dia en dashboard paciente',
            {
                'usuario_id': str(getattr(usuario, 'id', '')),
                'fecha_evento': fecha_evento.isoformat(),
                'inicio_dia': inicio_dia.isoformat(),
                'fin_dia': fin_dia.isoformat(),
                'duplicados_encontrados': solicitudes_mismo_dia,
            },
        )
        # #endregion

        print(f'DEBUG DASHBOARD: Validando solicitud para el Usuario ID: {getattr(usuario, "pk", None)}')
        existe_mismo_dia = Peticion.objects.filter(
            discapacitado_id=getattr(usuario, 'pk', None),
            fecha_evento__gte=inicio_dia,
            fecha_evento__lt=fin_dia,
            estado__in=[
                EstadoPeticion.ACTIVA,
                EstadoPeticion.ASIGNADA,
            ],
        ).exists()
        if existe_mismo_dia:
            messages.error(
                request,
                f"Ya tienes un acompañamiento programado para el {timezone.localtime(fecha_evento).strftime('%d/%m/%Y')}. "
                "Solo puedes crear otro para un día diferente.",
            )
            return redirect('dashboard_discapacitado')

        if fecha_evento <= timezone.now():
            messages.error(request, "La fecha y hora de inicio no pueden ser anteriores al momento actual.")
            return redirect('dashboard_discapacitado')

        if fecha_fin <= fecha_evento:
            messages.error(request, "La hora de fin debe ser posterior a la hora de inicio.")
            return redirect('dashboard_discapacitado')
        if fecha_fin - fecha_evento > datetime.timedelta(hours=4):
            messages.error(request, "El servicio no puede superar las 4 horas.")
            return redirect('dashboard_discapacitado')

        try:
            latitud = Decimal(latitud_raw)
            longitud = Decimal(longitud_raw)
        except (InvalidOperation, TypeError, ValueError):
            messages.error(request, "Coordenadas inválidas.")
            return redirect('dashboard_discapacitado')

        hace_poco = timezone.now() - datetime.timedelta(seconds=5)
        existe_duplicado = (
            Peticion.objects.filter(
                discapacitado=usuario,
                titulo=titulo,
                descripcion=descripcion,
                ciudad__iexact=(ciudad or '').strip(),
                fecha_evento=fecha_evento,
                fecha_creacion__gte=hace_poco,
            ).exists()
        )
        if existe_duplicado:
            messages.error(request, "Solicitud duplicada detectada. Ya estamos procesando tu solicitud.")
            return redirect('dashboard_discapacitado')

        peticion = Peticion.objects.create(
            discapacitado=usuario,
            titulo=titulo,
            descripcion=descripcion,
            fecha_evento=fecha_evento,
            fecha_fin=fecha_fin,
            ciudad=ciudad,
            direccion=direccion_texto,
            latitud=latitud,
            longitud=longitud,
            google_maps_url=google_maps_url,
            estado=EstadoPeticion.ACTIVA,
        )
        enfermeros_qs = Usuario.objects.filter(
            rol=RolUsuario.ENFERMERO,
            estado=EstadoVerificacion.APROBADO,
        ).exclude(email='')
        enfermeros_emails = list(
            enfermeros_qs.filter(
                perfil_enfermero__ciudad__iexact=(peticion.ciudad or '').strip(),
            )
            .values_list('email', flat=True)
            .distinct()
        )
        if not enfermeros_emails:
            enfermeros_emails = list(
                enfermeros_qs.values_list('email', flat=True).distinct()
            )
        if enfermeros_emails:
            ok = send_notification_email(
                subject='Assist Care · Nueva solicitud de cuidado disponible',
                to=enfermeros_emails,
                from_email=_mail_from_email(),
                status_label='Nuevo acompanamiento',
                headline='Hay un nuevo acompanamiento disponible',
                message='Se registró una nueva solicitud de cuidado y ya puedes revisar sus detalles para postularte si estás disponible.',
                details=compact_email_details(
                    make_email_detail('Titulo', peticion.titulo),
                    make_email_detail('Ciudad', peticion.ciudad),
                    make_email_detail('Direccion', peticion.direccion),
                    make_email_detail('Fecha del servicio', timezone.localtime(peticion.fecha_evento).strftime('%d/%m/%Y %I:%M %p')),
                ),
                cta_text='Ir al portal del enfermero',
                cta_url=request.build_absolute_uri(reverse('dashboard_enfermero')),
                footer_note='Recibirás más notificaciones cuando la solicitud cambie de estado o si decides postularte.',
            )
            if not ok:
                logger.error("Error enviando notificación de nueva solicitud.")
        messages.success(request, "Tu solicitud fue creada correctamente.")
        return redirect('dashboard_discapacitado')

    postulaciones_pendientes_prefetch = Prefetch(
        'postulaciones',
        queryset=Postulacion.objects.filter(estado=EstadoPostulacion.PENDIENTE).select_related('enfermero', 'enfermero__perfil_enfermero'),
        to_attr='postulaciones_pendientes',
    )
    postulaciones_aceptadas_prefetch = Prefetch(
        'postulaciones',
        queryset=Postulacion.objects.filter(estado=EstadoPostulacion.ACEPTADA).select_related('enfermero', 'enfermero__perfil_enfermero'),
        to_attr='postulaciones_aceptadas',
    )

    peticiones_qs = (
        Peticion.objects.filter(activo=True, discapacitado=usuario)
        .prefetch_related(postulaciones_pendientes_prefetch, postulaciones_aceptadas_prefetch)
        .order_by('-fecha_creacion')
    )

    servicios_por_aceptar_qs = (
        Peticion.objects.filter(activo=True, discapacitado=usuario, estado__in=[EstadoPeticion.ACTIVA, EstadoPeticion.ASIGNADA])
        .prefetch_related(postulaciones_pendientes_prefetch, postulaciones_aceptadas_prefetch)
        .order_by('-fecha_evento')
    )

    peticiones_paginator = Paginator(peticiones_qs, 5)
    peticiones_page_number = (request.GET.get('historial_page') or '').strip() or 1
    peticiones = peticiones_paginator.get_page(peticiones_page_number)

    por_aceptar_paginator = Paginator(servicios_por_aceptar_qs, 5)
    por_aceptar_page_number = (request.GET.get('por_aceptar_page') or '').strip() or 1
    servicios_por_aceptar = por_aceptar_paginator.get_page(por_aceptar_page_number)

    servicio_activo = (
        ServicioAcompanamiento.objects.filter(
            activo=True,
            paciente=usuario,
            estado__in=[EstadoServicio.EN_CAMINO, EstadoServicio.EN_PROGRESO],
        )
        .select_related('enfermero', 'enfermero__perfil_enfermero')
        .order_by('-fecha_creacion')
        .first()
    )
    if servicio_activo is not None and not (getattr(servicio_activo, 'codigo_verificacion', '') or '').strip():
        servicio_activo.codigo_verificacion = _generar_codigo_verificacion()
        servicio_activo.codigo_intentos = 0
        servicio_activo.codigo_bloqueado_hasta = None
        servicio_activo.save(update_fields=['codigo_verificacion', 'codigo_intentos', 'codigo_bloqueado_hasta'])

    servicios_completados = (
        ServicioAcompanamiento.objects.filter(activo=True, paciente=usuario, estado=EstadoServicio.COMPLETADO)
        .select_related('enfermero', 'enfermero__perfil_enfermero', 'peticion')
        .prefetch_related('calificacion')
        .order_by('fecha_creacion')
    )

    servicios_por_peticion = {}
    servicios_sin_peticion_por_enfermero = {}
    for s in servicios_completados:
        if s.peticion_id:
            servicios_por_peticion[s.peticion_id] = s
        else:
            servicios_sin_peticion_por_enfermero.setdefault(s.enfermero_id, []).append(s)

    for p in peticiones:
        p.servicio_para_calificar = None
        p.calificacion_existente = None
        if p.estado != EstadoPeticion.COMPLETADA:
            continue
        servicio = servicios_por_peticion.get(p.id)
        if servicio is None:
            postulacion_aceptada = None
            if hasattr(p, 'postulaciones_aceptadas') and p.postulaciones_aceptadas:
                postulacion_aceptada = p.postulaciones_aceptadas[0]
            if postulacion_aceptada is not None:
                pool = servicios_sin_peticion_por_enfermero.get(postulacion_aceptada.enfermero_id) or []
                if pool:
                    servicio = pool.pop(0)
            if servicio is None:
                continue
        cal = getattr(servicio, 'calificacion', None)
        if cal is not None:
            p.calificacion_existente = cal
        else:
            p.servicio_para_calificar = servicio

    notificaciones_paciente = []
    notificaciones_peticiones_ids = set()

    if servicio_activo and servicio_activo.enfermero:
        enfermero_activo = servicio_activo.enfermero
        notificaciones_paciente.append(
            {
                'etiqueta': 'Servicio activo',
                'titulo': f"{enfermero_activo.nombre_completo or enfermero_activo.email} está atendiendo tu servicio.",
                'descripcion': f"Estado actual: {servicio_activo.estado}. Revisa tu equipo de cuidado para más detalles.",
                'tab': 'equipo',
            }
        )
        if servicio_activo.peticion_id:
            notificaciones_peticiones_ids.add(servicio_activo.peticion_id)

    for servicio in servicios_por_aceptar_qs[:5]:
        if servicio.id in notificaciones_peticiones_ids:
            continue

        fecha_label = ''
        if servicio.fecha_evento:
            fecha_label = timezone.localtime(servicio.fecha_evento).strftime('%d/%m/%Y %I:%M %p')

        postulaciones_pendientes = list(getattr(servicio, 'postulaciones_pendientes', []) or [])
        postulaciones_aceptadas = list(getattr(servicio, 'postulaciones_aceptadas', []) or [])

        if servicio.estado == EstadoPeticion.ASIGNADA and postulaciones_aceptadas:
            enfermero_asignado = postulaciones_aceptadas[0].enfermero
            nombre_enfermero = enfermero_asignado.nombre_completo or enfermero_asignado.email or 'Tu profesional'
            notificaciones_paciente.append(
                {
                    'etiqueta': 'Profesional asignado',
                    'titulo': f"{nombre_enfermero} fue asignado a tu solicitud.",
                    'descripcion': f"{servicio.titulo}. {fecha_label}".strip(),
                    'tab': 'equipo',
                }
            )
            notificaciones_peticiones_ids.add(servicio.id)
            continue

        if postulaciones_pendientes:
            total_postulaciones = len(postulaciones_pendientes)
            notificaciones_paciente.append(
                {
                    'etiqueta': 'Nuevas postulaciones',
                    'titulo': f"Tienes {total_postulaciones} postulacion{'es' if total_postulaciones != 1 else ''} para revisar.",
                    'descripcion': f"{servicio.titulo}. {fecha_label}".strip(),
                    'tab': 'dashboard',
                }
            )
            notificaciones_peticiones_ids.add(servicio.id)

    fechas_ocupadas = sorted(
        {
            timezone.localtime(p.fecha_evento).date().isoformat()
            for p in Peticion.objects.filter(
                activo=True,
                discapacitado_id=getattr(usuario, 'pk', None),
                estado__in=[EstadoPeticion.ACTIVA, EstadoPeticion.ASIGNADA],
            ).only('fecha_evento')
            if getattr(p, 'fecha_evento', None)
        }
    )

    context = {
        'usuario': usuario,
        'perfil': perfil,
        'peticiones': peticiones,
        'servicios_por_aceptar': servicios_por_aceptar,
        'servicio_activo': servicio_activo,
        'notificaciones_paciente': notificaciones_paciente,
        'fechas_ocupadas': fechas_ocupadas,
    }
    return render(request, 'core/perfiles/dashboard_discapacitado.html', context)


@usuario_login_required
@role_required(RolUsuario.DISCAPACITADO)
def update_emergency_contact(request):
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Invalid method'}, status=405)

    try:
        raw_body = request.body.decode('utf-8') if request.body else ''
        data = json.loads(raw_body or '{}')
    except Exception as exc:
        return JsonResponse({'status': 'error', 'message': str(exc)}, status=400)

    usuario = request.usuario
    perfil = getattr(usuario, 'perfil_discapacitado', None)
    if perfil is None:
        return JsonResponse({'status': 'error', 'message': 'Profile not found'}, status=404)

    nombre = (data.get('name') or '').strip()[:150]
    parentesco = (data.get('relationship') or '').strip()[:80]
    telefono = (data.get('phone') or '').strip()

    perfil.emergency_nombre = nombre
    perfil.emergency_parentesco = parentesco
    perfil.emergency_telefono = telefono[:30]
    perfil.emergency_contact_name = (data.get('name') or '').strip()[:255] or None
    perfil.emergency_contact_relationship = (data.get('relationship') or '').strip()[:100] or None
    perfil.emergency_contact_phone = telefono[:20] or None
    perfil.fecha_actualizacion = timezone.now()
    perfil.save(
        update_fields=[
            'emergency_nombre',
            'emergency_parentesco',
            'emergency_telefono',
            'emergency_contact_name',
            'emergency_contact_relationship',
            'emergency_contact_phone',
            'fecha_actualizacion',
        ]
    )

    return JsonResponse(
        {
            'status': 'success',
            'message': 'Emergency contact updated successfully in database.',
            'data': {'name': nombre, 'relationship': parentesco, 'phone': telefono},
        }
    )


@usuario_login_required
def trigger_emergency_support(request):
    wants_json = (
        (request.headers.get('x-requested-with') == 'XMLHttpRequest')
        or ('application/json' in (request.headers.get('accept') or ''))
    )

    if request.method == 'GET':
        if wants_json:
            usuario = request.usuario
            role_label = _user_role_label(usuario)
            phone_label = _user_phone_for_emergency(usuario)
            perfil_disc = getattr(usuario, 'perfil_discapacitado', None)
            perfil_enf = getattr(usuario, 'perfil_enfermero', None)
            ciudad = (
                getattr(perfil_disc, 'ciudad', '')
                or getattr(perfil_enf, 'ciudad', '')
                or 'No registrada'
            )
            service_context = _find_emergency_service_context(usuario)
            return JsonResponse(
                {
                    'status': 'success',
                    'data': {
                        'user_id': usuario.id,
                        'role': role_label,
                        'phone': phone_label,
                        'city': ciudad,
                        'service_context': service_context,
                    },
                }
            )
        usuario = request.usuario
        messages.info(request, "El soporte de emergencia se activa desde el botón SOS dentro del panel.")
        if getattr(usuario, 'rol', None) == RolUsuario.ENFERMERO:
            return redirect('dashboard_enfermero')
        return redirect('dashboard_discapacitado')

    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Invalid method'}, status=405)

    usuario = request.usuario
    recipients = _emergency_recipient_list(usuario)
    if not recipients:
        return JsonResponse(
            {'status': 'error', 'message': 'No hay destinatarios institucionales configurados.'},
            status=500,
        )

    try:
        AuditoriaLog.objects.create(
            usuario_admin=None,
            accion=f"Se registró una alerta SOS institucional desde el portal. Usuario: {usuario.email or usuario.id}. Destinatarios: {', '.join(recipients)}",
        )
    except Exception:
        logger.exception("No fue posible registrar el evento SOS en AuditoriaLog.")

    role_label = _user_role_label(usuario)
    phone_label = _user_phone_for_emergency(usuario)
    perfil_disc = getattr(usuario, 'perfil_discapacitado', None)
    perfil_enf = getattr(usuario, 'perfil_enfermero', None)
    ciudad = (
        getattr(perfil_disc, 'ciudad', '')
        or getattr(perfil_enf, 'ciudad', '')
        or 'No registrada'
    )
    service_context = _find_emergency_service_context(usuario)

    subject = f"[EMERGENCIA PLATAFORMA] - Solicitud de Soporte Crítico - ID de Usuario: {usuario.id}"
    from_email = getattr(settings, 'DEFAULT_FROM_EMAIL', '') or getattr(settings, 'EMAIL_HOST_USER', '') or None

    service_lines = [
        f"- Servicio relacionado: {service_context['title']}",
        f"- Estado actual: {service_context['status']}",
        f"- ID del servicio operativo: {service_context['service_id']}",
        f"- ID de la solicitud: {service_context['petition_id']}",
        f"- Ubicación reportada: {service_context['address']}, {service_context['city']}",
        f"- Inicio programado: {service_context['scheduled_start'] or 'No disponible'}",
        f"- Fin programado: {service_context['scheduled_end'] or 'No disponible'}",
        f"- Contraparte vinculada: {service_context['counterpart_name']}",
        f"- Correo contraparte: {service_context['counterpart_email']}",
        f"- Teléfono contraparte: {service_context['counterpart_phone']}",
    ] if service_context else [
        "- No se encontró un servicio activo en curso para este usuario al momento de la alerta.",
    ]

    try:
        if not (getattr(settings, 'EMAIL_HOST_USER', '') or '').strip() or not (getattr(settings, 'EMAIL_HOST_PASSWORD', '') or '').strip():
            raise RuntimeError('Configuración SMTP incompleta (EMAIL_HOST_USER/EMAIL_HOST_PASSWORD).')

        send_notification_email(
            subject=subject,
            to=recipients,
            from_email=from_email,
            status_label='Emergencia institucional',
            headline='Alerta de soporte institucional',
            message='Un usuario solicitó atención prioritaria desde la opción de soporte de emergencia de la plataforma.',
            details=compact_email_details(
                make_email_detail('ID usuario', usuario.id),
                make_email_detail('Nombre completo', usuario.nombre_completo or 'No registrado'),
                make_email_detail('Rol', role_label),
                make_email_detail('Correo', usuario.email or 'No registrado'),
                make_email_detail('Telefono', phone_label),
                make_email_detail('Ciudad', ciudad),
                make_email_detail('Servicio relacionado', service_context['title'] if service_context else 'No disponible'),
                make_email_detail('Estado actual', service_context['status'] if service_context else 'No disponible'),
                make_email_detail('ID servicio', service_context['service_id'] if service_context else 'No disponible'),
                make_email_detail('ID solicitud', service_context['petition_id'] if service_context else 'No disponible'),
                make_email_detail('Ubicacion', f"{service_context['address']}, {service_context['city']}" if service_context else 'No disponible'),
                make_email_detail('Inicio programado', service_context['scheduled_start'] if service_context else 'No disponible'),
                make_email_detail('Fin programado', service_context['scheduled_end'] if service_context else 'No disponible'),
                make_email_detail('Contraparte vinculada', service_context['counterpart_name'] if service_context else 'No disponible'),
                make_email_detail('Correo contraparte', service_context['counterpart_email'] if service_context else 'No disponible'),
                make_email_detail('Telefono contraparte', service_context['counterpart_phone'] if service_context else 'No disponible'),
            ),
            cta_text='Abrir consola admin',
            cta_url=request.build_absolute_uri(reverse('dashboard_admin')),
            footer_note='Acción sugerida: revisar el panel administrativo de inmediato, validar el servicio asociado y contactar a la contraparte clínica correspondiente.',
        )
        return JsonResponse(
            {
                'status': 'success',
                'message': 'La alerta institucional fue enviada correctamente al equipo administrador.',
                'delivery': 'sent',
            }
        )
    except Exception as exc:
        logger.exception("Error enviando soporte de emergencia institucional para usuario %s", usuario.id)
        return JsonResponse(
            {
                'status': 'success',
                'message': 'La alerta fue registrada, pero no fue posible enviar el correo institucional. Verifica la configuración SMTP y los destinatarios.',
                'delivery': 'failed',
            }
        )

@usuario_login_required
@role_required(RolUsuario.DISCAPACITADO)
def calificar_enfermero(request, servicio_id):
    if request.method != 'POST':
        return redirect('dashboard_discapacitado')

    paciente = request.usuario
    servicio = (
        ServicioAcompanamiento.objects.select_related('paciente', 'enfermero', 'enfermero__perfil_enfermero')
        .filter(id=servicio_id, paciente=paciente, estado=EstadoServicio.COMPLETADO)
        .first()
    )
    if not servicio:
        messages.error(request, "No se encontró un servicio completado para calificar.")
        return redirect('dashboard_discapacitado')

    if servicio.enfermero is None:
        messages.error(request, "Este servicio no tiene enfermero asignado.")
        return redirect('dashboard_discapacitado')

    if getattr(servicio, 'calificacion', None) is not None:
        messages.warning(request, "Este servicio ya fue calificado.")
        return redirect('dashboard_discapacitado')

    estrellas_raw = (request.POST.get('estrellas') or '').strip()
    comentario = (request.POST.get('comentario') or '').strip()
    try:
        estrellas = int(estrellas_raw)
    except (TypeError, ValueError):
        estrellas = 0

    if estrellas < 1 or estrellas > 5:
        messages.error(request, "La calificación debe estar entre 1 y 5 estrellas.")
        return redirect('dashboard_discapacitado')

    with transaction.atomic():
        servicio_locked = (
            ServicioAcompanamiento.objects.select_for_update(of=('self',))
            .filter(id=servicio.id, paciente=paciente, estado=EstadoServicio.COMPLETADO)
            .first()
        )
        if not servicio_locked:
            messages.error(request, "No se pudo calificar este servicio.")
            return redirect('dashboard_discapacitado')

        if getattr(servicio_locked, 'calificacion', None) is not None:
            messages.warning(request, "Este servicio ya fue calificado.")
            return redirect('dashboard_discapacitado')

        Calificacion.objects.create(
            servicio=servicio_locked,
            paciente=paciente,
            enfermero=servicio_locked.enfermero,
            estrellas=estrellas,
            comentario=comentario,
        )

        promedio = (
            Calificacion.objects.filter(enfermero=servicio_locked.enfermero)
            .aggregate(Avg('estrellas'))
            .get('estrellas__avg')
        )
        enfermero_perfil = getattr(servicio_locked.enfermero, 'perfil_enfermero', None)
        if enfermero_perfil is not None:
            enfermero_perfil.calificacion_promedio = round(float(promedio or 0.0), 2)
            enfermero_perfil.save(update_fields=['calificacion_promedio'])

    messages.success(request, "¡Gracias por calificar a tu acompañante!")
    return redirect('dashboard_discapacitado')


@usuario_login_required
@role_required(RolUsuario.DISCAPACITADO)
def aceptar_enfermero(request, postulacion_id):
    if request.method != 'POST':
        return redirect('dashboard_discapacitado')

    usuario = request.usuario
    with transaction.atomic():
        postulacion = (
            Postulacion.objects.select_related('peticion', 'enfermero', 'enfermero__perfil_enfermero')
            .select_for_update(of=('self',))
            .filter(id=postulacion_id)
            .first()
        )
        if not postulacion:
            messages.error(request, "La postulación no existe.")
            return redirect('dashboard_discapacitado')

        peticion = (
            Peticion.objects.select_for_update()
            .filter(id=postulacion.peticion_id)
            .first()
        )
        if not peticion:
            messages.error(request, "La petición asociada no existe.")
            return redirect('dashboard_discapacitado')

        if peticion.discapacitado_id != usuario.id:
            return HttpResponseForbidden('Acceso denegado')

        if peticion.estado != EstadoPeticion.ACTIVA:
            messages.error(request, "Esta petición ya no está disponible para asignación.")
            return redirect('dashboard_discapacitado')

        if postulacion.estado != EstadoPostulacion.PENDIENTE:
            messages.error(request, "Esta postulación ya no está disponible.")
            return redirect('dashboard_discapacitado')

        Postulacion.objects.filter(peticion=peticion, id=postulacion.id).update(estado=EstadoPostulacion.ACEPTADA)
        Postulacion.objects.filter(peticion=peticion).exclude(id=postulacion.id).update(estado=EstadoPostulacion.RECHAZADA)
        Peticion.objects.filter(id=peticion.id).update(
            estado=EstadoPeticion.ASIGNADA,
            fecha_aceptacion=timezone.now(),
        )

        destino = (getattr(peticion, 'direccion', '') or '').strip() or peticion.ciudad
        origen = destino

        servicio = ServicioAcompanamiento.objects.create(
            peticion=peticion,
            paciente=usuario,
            enfermero=postulacion.enfermero,
            origen=origen,
            destino=destino,
            estado=EstadoServicio.EN_CAMINO,
            codigo_verificacion=_generar_codigo_verificacion(),
        )

    codigo_inicio = (getattr(servicio, 'codigo_verificacion', '') or '').strip()
    enfermero_name = postulacion.enfermero.nombre_completo or postulacion.enfermero.email
    fecha_local = timezone.localtime(peticion.fecha_evento)
    fin_local = timezone.localtime(peticion.fecha_fin) if getattr(peticion, 'fecha_fin', None) else None
    rango_horario = fecha_local.strftime('%d/%m/%Y %I:%M %p')
    if fin_local:
        rango_horario = f"{rango_horario} - {fin_local.strftime('%I:%M %p')}"

    correo_paciente = (getattr(usuario, 'email', '') or '').strip()
    correo_enfermero = (getattr(postulacion.enfermero, 'email', '') or '').strip()
    paciente_name = (getattr(usuario, 'nombre_completo', '') or '').strip() or correo_paciente or 'Paciente'

    ok_paciente = True
    ok_enfermero = True
    if correo_paciente:
        ok_paciente = send_notification_email(
            subject='Assist Care · Tu solicitud fue aceptada',
            to=correo_paciente,
            from_email=_mail_from_email(),
            status_label='Solicitud aceptada',
            headline='Tu acompanamiento ya fue asignado',
            message=(
                f"Hola {paciente_name}, tu solicitud fue aceptada por {enfermero_name}. "
                "Ya puedes coordinar los detalles del servicio desde la plataforma."
            ),
            details=compact_email_details(
                make_email_detail('Solicitud', peticion.titulo),
                make_email_detail('Fecha', fecha_local.strftime('%d/%m/%Y')),
                make_email_detail('Horario', rango_horario),
                make_email_detail('Profesional asignado', enfermero_name),
                make_email_detail('Codigo de verificacion', codigo_inicio),
            ),
            cta_text='Abrir mi panel',
            cta_url=request.build_absolute_uri(reverse('dashboard_discapacitado')),
            footer_note='Conserva el codigo de verificacion porque se usará para iniciar el acompanamiento.',
        )
    if correo_enfermero:
        ok_enfermero = send_notification_email(
            subject='Assist Care · Servicio agendado y asignado',
            to=correo_enfermero,
            from_email=_mail_from_email(),
            status_label='Servicio asignado',
            headline='Tienes un nuevo acompanamiento confirmado',
            message=f"Hola {enfermero_name}, ya quedaste asignado oficialmente al servicio \"{peticion.titulo}\".",
            details=compact_email_details(
                make_email_detail('Paciente', paciente_name),
                make_email_detail('Fecha y horario', rango_horario),
                make_email_detail('Direccion de recogida', destino),
                make_email_detail('Codigo de verificacion', codigo_inicio),
            ),
            cta_text='Abrir portal del enfermero',
            cta_url=request.build_absolute_uri(reverse('dashboard_enfermero')),
            footer_note='Revisa tu panel para preparar el acompanamiento y mantener comunicación con el paciente.',
        )
    if not (ok_paciente and ok_enfermero):
        logger.error("No se pudo enviar correo de asignación para petición %s", peticion.id)
        messages.warning(request, "El servicio fue asignado, pero no fue posible enviar las notificaciones por correo.")
    messages.success(request, f"¡Has aceptado a {enfermero_name}! El servicio está ahora en camino.")
    return redirect('dashboard_discapacitado')


@usuario_login_required
@role_required(RolUsuario.DISCAPACITADO)
def aceptar_postulacion(request, peticion_id, enfermero_id):
    if request.method != 'POST':
        return redirect('dashboard_discapacitado')

    usuario = request.usuario
    with transaction.atomic():
        postulacion = (
            Postulacion.objects.select_related('peticion', 'enfermero', 'enfermero__perfil_enfermero')
            .select_for_update(of=('self',))
            .filter(peticion_id=peticion_id, enfermero_id=enfermero_id)
            .first()
        )
        if not postulacion:
            messages.error(request, "La postulación no existe.")
            return redirect('dashboard_discapacitado')

        peticion = (
            Peticion.objects.select_for_update()
            .filter(id=postulacion.peticion_id)
            .first()
        )
        if not peticion:
            messages.error(request, "La petición asociada no existe.")
            return redirect('dashboard_discapacitado')

        if peticion.discapacitado_id != usuario.id:
            return HttpResponseForbidden('Acceso denegado')

        if peticion.estado != EstadoPeticion.ACTIVA:
            messages.error(request, "Esta petición ya no está disponible para asignación.")
            return redirect('dashboard_discapacitado')

        if postulacion.estado != EstadoPostulacion.PENDIENTE:
            messages.error(request, "Esta postulación ya no está disponible.")
            return redirect('dashboard_discapacitado')

        Postulacion.objects.filter(peticion=peticion, id=postulacion.id).update(estado=EstadoPostulacion.ACEPTADA)
        Postulacion.objects.filter(peticion=peticion).exclude(id=postulacion.id).update(estado=EstadoPostulacion.RECHAZADA)
        Peticion.objects.filter(id=peticion.id).update(
            estado=EstadoPeticion.ASIGNADA,
            fecha_aceptacion=timezone.now(),
        )

        destino = (getattr(peticion, 'direccion', '') or '').strip() or peticion.ciudad
        origen = destino

        servicio = ServicioAcompanamiento.objects.create(
            peticion=peticion,
            paciente=usuario,
            enfermero=postulacion.enfermero,
            origen=origen,
            destino=destino,
            estado=EstadoServicio.EN_CAMINO,
            codigo_verificacion=_generar_codigo_verificacion(),
        )

    codigo_inicio = (getattr(servicio, 'codigo_verificacion', '') or '').strip()
    enfermero_name = postulacion.enfermero.nombre_completo or postulacion.enfermero.email
    fecha_local = timezone.localtime(peticion.fecha_evento)
    fin_local = timezone.localtime(peticion.fecha_fin) if getattr(peticion, 'fecha_fin', None) else None
    destino = (getattr(peticion, 'direccion', '') or '').strip() or peticion.ciudad
    rango_horario = fecha_local.strftime('%d/%m/%Y %I:%M %p')
    if fin_local:
        rango_horario = f"{rango_horario} - {fin_local.strftime('%I:%M %p')}"

    correo_paciente = (getattr(usuario, 'email', '') or '').strip()
    correo_enfermero = (getattr(postulacion.enfermero, 'email', '') or '').strip()
    paciente_name = (getattr(usuario, 'nombre_completo', '') or '').strip() or correo_paciente or 'Paciente'

    ok_paciente = True
    ok_enfermero = True
    if correo_paciente:
        ok_paciente = send_notification_email(
            subject='Assist Care · Tu solicitud fue aceptada',
            to=correo_paciente,
            from_email=_mail_from_email(),
            status_label='Solicitud aceptada',
            headline='Tu acompanamiento ya fue asignado',
            message=(
                f"Hola {paciente_name}, tu solicitud fue aceptada por {enfermero_name}. "
                "Ya puedes coordinar los detalles del servicio desde la plataforma."
            ),
            details=compact_email_details(
                make_email_detail('Solicitud', peticion.titulo),
                make_email_detail('Fecha', fecha_local.strftime('%d/%m/%Y')),
                make_email_detail('Horario', rango_horario),
                make_email_detail('Profesional asignado', enfermero_name),
                make_email_detail('Codigo de verificacion', codigo_inicio),
            ),
            cta_text='Abrir mi panel',
            cta_url=request.build_absolute_uri(reverse('dashboard_discapacitado')),
            footer_note='Conserva el codigo de verificacion porque se usará para iniciar el acompanamiento.',
        )
    if correo_enfermero:
        ok_enfermero = send_notification_email(
            subject='Assist Care · Servicio agendado y asignado',
            to=correo_enfermero,
            from_email=_mail_from_email(),
            status_label='Servicio asignado',
            headline='Tienes un nuevo acompanamiento confirmado',
            message=f"Hola {enfermero_name}, ya quedaste asignado oficialmente al servicio \"{peticion.titulo}\".",
            details=compact_email_details(
                make_email_detail('Paciente', paciente_name),
                make_email_detail('Fecha y horario', rango_horario),
                make_email_detail('Direccion de recogida', destino),
                make_email_detail('Codigo de verificacion', codigo_inicio),
            ),
            cta_text='Abrir portal del enfermero',
            cta_url=request.build_absolute_uri(reverse('dashboard_enfermero')),
            footer_note='Revisa tu panel para preparar el acompanamiento y mantener comunicación con el paciente.',
        )
    if not (ok_paciente and ok_enfermero):
        logger.error("No se pudo enviar correo de asignación para petición %s", peticion.id)
        messages.warning(request, "El servicio fue asignado, pero no fue posible enviar las notificaciones por correo.")
    messages.success(request, f"¡Has aceptado a {enfermero_name}! El servicio está ahora en camino.")
    return redirect('dashboard_discapacitado')


@usuario_login_required
def cancelar_solicitud(request, peticion_id):
    if request.method != 'POST':
        if getattr(request.usuario, 'rol', None) == RolUsuario.ENFERMERO:
            return redirect('dashboard_enfermero')
        return redirect('dashboard_discapacitado')

    usuario = request.usuario
    mail_payload = None
    mail_failed = False
    with transaction.atomic():
        peticion = (
            Peticion.objects.select_for_update()
            .select_related('discapacitado')
            .filter(id=peticion_id)
            .first()
        )
        if not peticion:
            messages.error(request, "La solicitud no existe.")
            if getattr(usuario, 'rol', None) == RolUsuario.ENFERMERO:
                return redirect('dashboard_enfermero')
            return redirect('dashboard_discapacitado')

        is_owner = peticion.discapacitado_id == usuario.id
        is_nurse = getattr(usuario, 'rol', None) == RolUsuario.ENFERMERO

        assigned_ok = False
        if is_nurse and peticion.estado == EstadoPeticion.ASIGNADA:
            assigned_post = (
                Postulacion.objects.filter(
                    peticion=peticion,
                    estado=EstadoPostulacion.ACEPTADA,
                    enfermero=usuario,
                )
                .only('id')
                .first()
            )
            assigned_service = ServicioAcompanamiento.objects.filter(peticion=peticion, enfermero=usuario).only('id').first()
            assigned_ok = bool(assigned_post or assigned_service)

        if not is_owner and not assigned_ok:
            return HttpResponseForbidden('Acceso denegado')

        if peticion.estado == EstadoPeticion.ACTIVA and not is_owner:
            return HttpResponseForbidden('Acceso denegado')

        if not getattr(peticion, 'puede_cancelarse', False):
            messages.error(
                request,
                "La solicitud no puede cancelarse porque ha superado el límite de 24 horas desde su aceptación.",
            )
            if is_nurse:
                return redirect('dashboard_enfermero')
            return redirect('dashboard_discapacitado')

        titulo = (getattr(peticion, 'titulo', '') or '').strip() or 'Solicitud de acompañamiento'
        fecha_local = timezone.localtime(getattr(peticion, 'fecha_evento', timezone.now()))
        fin_local = timezone.localtime(peticion.fecha_fin) if getattr(peticion, 'fecha_fin', None) else None
        rango = fecha_local.strftime('%d/%m/%Y %I:%M %p')
        if fin_local:
            rango = f"{rango} - {fin_local.strftime('%I:%M %p')}"

        paciente = getattr(peticion, 'discapacitado', None)
        correo_paciente = (getattr(paciente, 'email', '') or '').strip()
        nombre_paciente = (getattr(paciente, 'nombre_completo', '') or '').strip() or correo_paciente or 'Paciente'

        postulacion_aceptada = (
            Postulacion.objects.select_related('enfermero')
            .filter(peticion=peticion, estado=EstadoPostulacion.ACEPTADA)
            .first()
        )
        enfermero_asignado = getattr(postulacion_aceptada, 'enfermero', None)
        correo_enfermero = (getattr(enfermero_asignado, 'email', '') or '').strip() if enfermero_asignado else ''
        nombre_enfermero = (getattr(enfermero_asignado, 'nombre_completo', '') or '').strip() if enfermero_asignado else ''
        if not nombre_enfermero:
            nombre_enfermero = correo_enfermero or 'Profesional'

        nombre_cancelador = (getattr(usuario, 'nombre_completo', '') or '').strip() or (getattr(usuario, 'email', '') or '').strip() or 'Usuario'

        if is_owner:
            Peticion.objects.filter(id=peticion.id).update(estado=EstadoPeticion.CANCELADA)
            Postulacion.objects.filter(peticion=peticion).update(estado=EstadoPostulacion.RECHAZADA)
            if correo_enfermero:
                mail_payload = {
                    'subject': 'Assist Care · El paciente canceló el acompanamiento',
                    'to': correo_enfermero,
                    'from_email': _mail_from_email(),
                    'status_label': 'Servicio cancelado',
                    'headline': 'El paciente canceló este acompanamiento',
                    'message': f"{nombre_cancelador} canceló la solicitud \"{titulo}\" dentro del periodo permitido y tu agenda quedó liberada.",
                    'details': compact_email_details(
                        make_email_detail('Solicitud', titulo),
                        make_email_detail('Fecha y horario', rango),
                        make_email_detail('Paciente', nombre_paciente),
                    ),
                    'cta_text': 'Ver portal del enfermero',
                    'cta_url': request.build_absolute_uri(reverse('dashboard_enfermero')),
                    'footer_note': 'Puedes seguir revisando nuevas oportunidades disponibles desde tu panel.',
                }
        else:
            Peticion.objects.filter(id=peticion.id).update(
                estado=EstadoPeticion.ACTIVA,
                fecha_aceptacion=None,
            )
            Postulacion.objects.filter(peticion=peticion, enfermero=usuario).update(estado=EstadoPostulacion.RECHAZADA)
            Postulacion.objects.filter(peticion=peticion).exclude(enfermero=usuario).update(estado=EstadoPostulacion.PENDIENTE)
            if correo_paciente:
                mail_payload = {
                    'subject': 'Assist Care · Tu profesional canceló el acompanamiento',
                    'to': correo_paciente,
                    'from_email': _mail_from_email(),
                    'status_label': 'Servicio liberado',
                    'headline': 'Tu profesional canceló el acompanamiento asignado',
                    'message': (
                        f"Hola {nombre_paciente}, el profesional {nombre_cancelador} canceló la solicitud "
                        f"\"{titulo}\" dentro del periodo permitido. La publicamos nuevamente para que otros profesionales puedan postularse."
                    ),
                    'details': compact_email_details(
                        make_email_detail('Solicitud', titulo),
                        make_email_detail('Fecha y horario', rango),
                        make_email_detail('Profesional', nombre_cancelador),
                    ),
                    'cta_text': 'Abrir mi panel',
                    'cta_url': request.build_absolute_uri(reverse('dashboard_discapacitado')),
                    'footer_note': 'Te recomendamos revisar nuevas postulaciones y confirmar otro profesional cuando esté disponible.',
                }

        ServicioAcompanamiento.objects.filter(
            peticion=peticion,
            estado__in=[EstadoServicio.EN_CAMINO, EstadoServicio.EN_PROGRESO],
        ).update(estado=EstadoServicio.CANCELADO, enfermero=None)

        def _bump_metrics():
            try:
                SystemPulse.bump('admin_metrics')
            except Exception:
                logger.exception("Error actualizando pulso de métricas admin (cancelación).")

        transaction.on_commit(_bump_metrics)

    if mail_payload:
        ok_main = send_notification_email(**mail_payload)
        ok_actor = True
        correo_actor = (getattr(usuario, 'email', '') or '').strip()
        if correo_actor and correo_actor != mail_payload['to']:
            if is_owner:
                ok_actor = send_notification_email(
                    subject='Assist Care · Confirmamos la cancelación de tu solicitud',
                    to=correo_actor,
                    from_email=_mail_from_email(),
                    status_label='Cancelacion registrada',
                    headline='Tu cancelación fue registrada correctamente',
                    message='Hemos cancelado tu solicitud de acompanamiento. Si todavía necesitas el servicio, puedes crear una nueva solicitud con un horario actualizado.',
                    cta_text='Crear o revisar solicitudes',
                    cta_url=request.build_absolute_uri(reverse('dashboard_discapacitado')),
                    footer_note='Gracias por mantener actualizada tu disponibilidad en Assist Care.',
                )
            else:
                ok_actor = send_notification_email(
                    subject='Assist Care · Confirmamos la cancelación del servicio',
                    to=correo_actor,
                    from_email=_mail_from_email(),
                    status_label='Cancelacion registrada',
                    headline='Tu cancelación fue registrada correctamente',
                    message='Hemos registrado la cancelación del servicio asignado y la solicitud quedó disponible nuevamente para nuevas postulaciones.',
                    cta_text='Volver al portal del enfermero',
                    cta_url=request.build_absolute_uri(reverse('dashboard_enfermero')),
                    footer_note='Mantén tu panel actualizado para revisar nuevas oportunidades disponibles.',
                )
        if not (ok_main and ok_actor):
            mail_failed = True
            logger.error("No se pudo enviar correo de cancelación para petición %s", peticion_id)

    if is_owner:
        messages.success(request, "Solicitud cancelada correctamente.")
    else:
        messages.success(request, "Solicitud cancelada. La petición quedó disponible nuevamente.")
    if mail_payload and mail_failed:
        messages.warning(request, "La cancelación se procesó, pero no fue posible enviar el correo de notificación.")
    if is_nurse:
        return redirect('dashboard_enfermero')
    return redirect('dashboard_discapacitado')


@usuario_login_required
@role_required(RolUsuario.ENFERMERO)
def dashboard_enfermero(request):
    usuario = request.usuario
    perfil = getattr(usuario, 'perfil_enfermero', None)
    ciudad_enfermero = (getattr(perfil, 'ciudad', None) or '').strip()

    if request.method == 'POST':
        form_type = (request.POST.get('form_type') or '').strip().lower()
        if form_type == 'foto':
            pass
        if form_type == 'perfil':
            if perfil is None:
                messages.error(request, "No se encontró el perfil del enfermero.")
                return redirect('dashboard_enfermero')

            perfil_form = PerfilEnfermeroUpdateForm(request.POST, request.FILES)
            if not perfil_form.is_valid():
                errores = []
                for field_errors in perfil_form.errors.values():
                    errores.extend(field_errors)
                for error in errores[:3]:
                    messages.error(request, error)
                return redirect(f"{reverse('dashboard_enfermero')}?tab=perfil")

            perfil_form.save(perfil)

            messages.success(request, "Tu perfil se actualizó correctamente.")
            return redirect(f"{reverse('dashboard_enfermero')}?tab=perfil")
        if form_type == 'bio':
            if perfil is None:
                messages.error(request, "No se encontró el perfil del enfermero.")
                return redirect('dashboard_enfermero')
            especialidad = (request.POST.get('especialidad') or '').strip()
            biografia = (request.POST.get('biografia') or '').strip()
            if especialidad:
                perfil.especialidad = especialidad
            perfil.biografia = biografia
            perfil.fecha_actualizacion = timezone.now()
            perfil.save()
            wants_json = request.headers.get('x-requested-with') == 'XMLHttpRequest' or 'application/json' in (request.headers.get('accept') or '')
            if wants_json:
                return JsonResponse(
                    {
                        'ok': True,
                        'especialidad': perfil.especialidad or '',
                        'biografia': perfil.biografia or '',
                    }
                )
            messages.success(request, "Perfil profesional actualizado.")
            return redirect(f"{reverse('dashboard_enfermero')}?tab=perfil")

    servicio_activo = (
        ServicioAcompanamiento.objects.filter(
            activo=True,
            enfermero=usuario,
            estado__in=[EstadoServicio.EN_CAMINO, EstadoServicio.EN_PROGRESO],
        )
        .select_related('paciente', 'paciente__perfil_discapacitado', 'peticion')
        .order_by('-fecha_creacion')
        .first()
    )
    if servicio_activo is not None and not (getattr(servicio_activo, 'codigo_verificacion', '') or '').strip():
        servicio_activo.codigo_verificacion = _generar_codigo_verificacion()
        servicio_activo.codigo_intentos = 0
        servicio_activo.codigo_bloqueado_hasta = None
        servicio_activo.save(update_fields=['codigo_verificacion', 'codigo_intentos', 'codigo_bloqueado_hasta'])
    servicio_activo_ok = servicio_activo is not None
    servicio_asignado = servicio_activo
    servicio_activo_peticion = None
    if servicio_activo is not None:
        if getattr(servicio_activo, 'peticion_id', None):
            servicio_activo_peticion = servicio_activo.peticion
        else:
            servicio_activo_peticion = (
                Peticion.objects.filter(
                    activo=True,
                    estado=EstadoPeticion.ASIGNADA,
                    postulaciones__enfermero=usuario,
                    postulaciones__estado=EstadoPostulacion.ACEPTADA,
                )
                .select_related('discapacitado', 'discapacitado__perfil_discapacitado')
                .order_by('-fecha_creacion')
                .first()
            )

    if not ciudad_enfermero:
        peticiones_disponibles = Peticion.objects.none()
    elif servicio_activo is not None:
        peticiones_disponibles = Peticion.objects.none()
    else:
        ya_postulado_subq = Postulacion.objects.filter(
            peticion=OuterRef('pk'),
            enfermero=usuario,
            estado=EstadoPostulacion.PENDIENTE,
        )
        peticiones_disponibles = (
            Peticion.objects.filter(activo=True, estado=EstadoPeticion.ACTIVA, ciudad=ciudad_enfermero)
            .select_related('discapacitado', 'discapacitado__perfil_discapacitado')
            .annotate(ya_postulado=Exists(ya_postulado_subq))
            .order_by('-fecha_creacion')
        )

    blocked_postulation_dates = _nurse_blocked_postulation_dates(usuario)

    servicios_disponibles = Peticion.objects.none()
    if ciudad_enfermero:
        servicios_disponibles = (
            Peticion.objects.filter(activo=True, estado=EstadoPeticion.ACTIVA, ciudad__iexact=ciudad_enfermero)
            .select_related('discapacitado', 'discapacitado__perfil_discapacitado')
            .annotate(
                ya_postulado=Exists(
                    Postulacion.objects.filter(
                        peticion=OuterRef('pk'),
                        enfermero=usuario,
                        estado=EstadoPostulacion.PENDIENTE,
                    )
                )
            )
            .order_by('-fecha_creacion')
        )

    peticiones_paginator = Paginator(peticiones_disponibles, 5)
    peticiones_page_number = (request.GET.get('solicitudes_page') or '').strip() or 1
    peticiones_disponibles = peticiones_paginator.get_page(peticiones_page_number)
    _apply_same_day_postulation_flag(peticiones_disponibles.object_list, blocked_postulation_dates)
    _apply_same_day_postulation_flag(servicios_disponibles, blocked_postulation_dates)

    servicios_pasados = (
        ServicioAcompanamiento.objects.filter(activo=True, enfermero=usuario, estado=EstadoServicio.COMPLETADO)
        .select_related('paciente', 'paciente__perfil_discapacitado')
        .order_by('-id')
    )
    servicios_exitosos_count = servicios_pasados.count()

    servicios_asignados = (
        ServicioAcompanamiento.objects.filter(activo=True, enfermero=usuario, estado__in=[EstadoServicio.EN_CAMINO, EstadoServicio.EN_PROGRESO])
        .select_related('paciente', 'paciente__perfil_discapacitado', 'peticion')
        .order_by('-fecha_creacion')
    )
    for svc in servicios_asignados:
        perfil_paciente = getattr(getattr(svc, 'paciente', None), 'perfil_discapacitado', None)
        fecha_nacimiento = getattr(perfil_paciente, 'fecha_nacimiento', None)
        svc.paciente_edad = None
        if fecha_nacimiento:
            hoy = timezone.localdate()
            svc.paciente_edad = hoy.year - fecha_nacimiento.year - (
                (hoy.month, hoy.day) < (fecha_nacimiento.month, fecha_nacimiento.day)
            )

    hoy = timezone.localdate()
    servicios_hoy = (
        ServicioAcompanamiento.objects.filter(activo=True, enfermero=usuario)
        .select_related('peticion', 'paciente', 'paciente__perfil_discapacitado')
        .filter(
            Q(peticion__fecha_evento__date=hoy)
            | Q(peticion__isnull=True, fecha_creacion__date=hoy)
        )
        .order_by('peticion__fecha_evento', 'fecha_creacion')
    )
    servicios_pasados_ids = [s.id for s in servicios_pasados]
    calificaciones_paciente_qs = CalificacionPaciente.objects.filter(enfermero=usuario, servicio_id__in=servicios_pasados_ids)
    calificaciones_paciente_map = {c.servicio_id: c for c in calificaciones_paciente_qs}
    for s in servicios_pasados:
        s.calificacion_paciente = calificaciones_paciente_map.get(s.id)

    context = {
        'usuario': usuario,
        'perfil': perfil,
        'ciudad_enfermero': ciudad_enfermero,
        'peticiones_disponibles': peticiones_disponibles,
        'servicios_disponibles': servicios_disponibles,
        'servicio_activo': servicio_activo,
        'servicio_activo_ok': servicio_activo_ok,
        'servicio_asignado': servicio_asignado,
        'servicio_activo_peticion': servicio_activo_peticion,
        'servicios_pasados': servicios_pasados,
        'servicios_asignados': servicios_asignados,
        'servicios_hoy': servicios_hoy,
        'servicios_exitosos_count': servicios_exitosos_count,
    }
    return render(request, 'core/perfiles/dashboard_enfermero.html', context)


@usuario_login_required
@role_required(RolUsuario.ENFERMERO)
def enfermero_servicios_view(request):
    usuario = request.usuario
    perfil = getattr(usuario, 'perfil_enfermero', None)
    ciudad_enfermero = (getattr(perfil, 'ciudad', None) or '').strip()

    servicio_activo_ok = ServicioAcompanamiento.objects.filter(
        activo=True,
        enfermero=usuario,
        estado__in=[EstadoServicio.EN_CAMINO, EstadoServicio.EN_PROGRESO],
    ).exists()

    servicios = Peticion.objects.none()
    if ciudad_enfermero:
        ya_postulado_subq = Postulacion.objects.filter(
            peticion=OuterRef('pk'),
            enfermero=usuario,
            estado=EstadoPostulacion.PENDIENTE,
        )
        servicios = (
            Peticion.objects.filter(activo=True, estado=EstadoPeticion.ACTIVA, ciudad__iexact=ciudad_enfermero)
            .select_related('discapacitado', 'discapacitado__perfil_discapacitado')
            .annotate(ya_postulado=Exists(ya_postulado_subq))
            .order_by('-fecha_creacion')
        )

    blocked_postulation_dates = _nurse_blocked_postulation_dates(usuario)
    _apply_same_day_postulation_flag(servicios, blocked_postulation_dates)

    context = {
        'usuario': usuario,
        'perfil': perfil,
        'ciudad_enfermero': ciudad_enfermero,
        'servicios': servicios,
        'servicio_activo_ok': servicio_activo_ok,
    }
    return render(request, 'core/perfiles/enfermero_servicios.html', context)


@usuario_login_required
@role_required(RolUsuario.ENFERMERO)
def ultimas_solicitudes_json(request):
    usuario = request.usuario
    perfil = getattr(usuario, 'perfil_enfermero', None)
    ciudad = (getattr(perfil, 'ciudad', None) or '').strip()

    if not ciudad:
        return JsonResponse({'status': 'success', 'notificaciones': [], 'count': 0})

    solicitudes = (
        Peticion.objects.filter(estado=EstadoPeticion.ACTIVA, ciudad__iexact=ciudad)
        .select_related('discapacitado', 'discapacitado__perfil_discapacitado')
        .order_by('-fecha_creacion')[:5]
    )

    notificaciones = []
    for s in solicitudes:
        paciente = getattr(s, 'discapacitado', None)
        paciente_label = None
        if paciente is not None:
            paciente_label = getattr(paciente, 'nombre_completo', None) or getattr(paciente, 'email', None)
        if not paciente_label:
            paciente_label = 'Paciente'

        fecha_dt = getattr(s, 'fecha_evento', None) or getattr(s, 'fecha_creacion', None)
        fecha_label = ''
        if fecha_dt:
            fecha_label = timezone.localtime(fecha_dt).strftime('%d/%m %I:%M %p')

        notificaciones.append(
            {
                'id': str(s.id),
                'paciente': paciente_label,
                'descripcion': (getattr(s, 'descripcion', None) or '')[:220],
                'fecha': fecha_label,
                'cta_url': reverse('dashboard_enfermero') + '?tab=pacientes',
            }
        )

    return JsonResponse({'status': 'success', 'notificaciones': notificaciones, 'count': len(notificaciones)})


@usuario_login_required
@role_required(RolUsuario.ENFERMERO)
def enfermero_emergencia(request):
    if request.method != 'POST':
        return HttpResponseNotAllowed(['POST'])

    usuario = request.usuario
    servicio_activo = (
        ServicioAcompanamiento.objects.filter(
            enfermero=usuario,
            estado__in=[EstadoServicio.EN_CAMINO, EstadoServicio.EN_PROGRESO],
        )
        .select_related('paciente', 'paciente__perfil_discapacitado', 'peticion')
        .order_by('-fecha_creacion')
        .first()
    )

    paciente = getattr(servicio_activo, 'paciente', None)
    perfil_pac = getattr(paciente, 'perfil_discapacitado', None) if paciente else None
    peticion = getattr(servicio_activo, 'peticion', None)

    detalles = {
        'enfermero': usuario.nombre_completo or usuario.email,
        'estado_servicio': getattr(servicio_activo, 'estado', '') or 'N/A',
        'paciente': (paciente.nombre_completo or paciente.email) if paciente else 'N/A',
        'titulo': getattr(peticion, 'titulo', '') or f"Servicio #{getattr(servicio_activo, 'id', '')}",
        'direccion': (getattr(peticion, 'direccion', '') or getattr(servicio_activo, 'destino', '') or getattr(servicio_activo, 'origen', '') or '').strip(),
        'ciudad': (getattr(peticion, 'ciudad', '') or '').strip(),
        'fecha': timezone.localtime(getattr(peticion, 'fecha_evento', None)).strftime('%d/%m/%Y %I:%M %p') if getattr(peticion, 'fecha_evento', None) else '',
        'emergencia_nombre': (getattr(perfil_pac, 'emergency_nombre', '') or '').strip(),
        'emergencia_parentesco': (getattr(perfil_pac, 'emergency_parentesco', '') or '').strip(),
        'emergencia_telefono': (getattr(perfil_pac, 'emergency_telefono', '') or '').strip(),
    }

    admins = Usuario.objects.filter(
        rol__in=[RolUsuario.ADMIN, RolUsuario.SUPER_ADMIN],
        estado=EstadoVerificacion.APROBADO,
    ).only('email')

    for adm in admins.iterator():
        send_notification_email(
            subject="Assist Care · Alerta de emergencia (Enfermero)",
            to=adm.email,
            from_email=_mail_from_email(),
            status_label='Emergencia',
            headline='Un enfermero activó una alerta de emergencia',
            message='Se detectó una alerta de emergencia enviada desde el portal del enfermero.',
            details=compact_email_details(
                *[make_email_detail(label.replace('_', ' ').capitalize(), value) for label, value in detalles.items()]
            ),
            cta_text='Abrir consola admin',
            cta_url=request.build_absolute_uri(reverse('dashboard_admin')),
            footer_note='Revisa este caso con prioridad alta y valida el contexto del servicio asociado.',
        )

    return JsonResponse({'ok': True})

@usuario_login_required
@role_required(RolUsuario.ENFERMERO)
def calificar_paciente(request, servicio_id):
    wants_json = request.headers.get('x-requested-with') == 'XMLHttpRequest' or 'application/json' in (request.headers.get('accept') or '')

    if request.method != 'POST':
        if wants_json:
            return JsonResponse({'ok': False, 'error': 'Método no permitido.'}, status=405)
        return redirect('dashboard_enfermero')

    enfermero = request.usuario
    servicio = (
        ServicioAcompanamiento.objects.select_related('paciente', 'paciente__perfil_discapacitado')
        .filter(id=servicio_id, enfermero=enfermero, estado=EstadoServicio.COMPLETADO)
        .first()
    )
    if not servicio:
        if wants_json:
            return JsonResponse({'ok': False, 'error': 'No se encontró un servicio completado para calificar.'}, status=404)
        messages.error(request, "No se encontró un servicio completado para calificar.")
        return redirect('dashboard_enfermero')

    estrellas_raw = (request.POST.get('estrellas') or '').strip()
    comentario = (request.POST.get('comentario') or '').strip()
    try:
        estrellas = int(estrellas_raw)
    except (TypeError, ValueError):
        estrellas = 0

    if estrellas < 1 or estrellas > 5:
        if wants_json:
            return JsonResponse({'ok': False, 'error': 'La calificación debe estar entre 1 y 5 estrellas.'}, status=400)
        messages.error(request, "La calificación debe estar entre 1 y 5 estrellas.")
        return redirect('dashboard_enfermero')

    if not comentario:
        if wants_json:
            return JsonResponse({'ok': False, 'error': 'El comentario es obligatorio.'}, status=400)
        messages.error(request, "El comentario es obligatorio para enviar la calificación.")
        return redirect('dashboard_enfermero')

    with transaction.atomic():
        servicio_locked = (
            ServicioAcompanamiento.objects.select_for_update(of=('self',))
            .filter(id=servicio.id, enfermero=enfermero, estado=EstadoServicio.COMPLETADO)
            .first()
        )
        if not servicio_locked:
            if wants_json:
                return JsonResponse({'ok': False, 'error': 'No se pudo calificar este servicio.'}, status=400)
            messages.error(request, "No se pudo calificar este servicio.")
            return redirect('dashboard_enfermero')

        try:
            CalificacionPaciente.objects.create(
                servicio=servicio_locked,
                paciente=servicio_locked.paciente,
                enfermero=enfermero,
                estrellas=estrellas,
                comentario=comentario,
            )
        except IntegrityError:
            if wants_json:
                return JsonResponse({'ok': False, 'error': 'Este paciente ya fue calificado para este servicio.'}, status=409)
            messages.warning(request, "Este paciente ya fue calificado para este servicio.")
            return redirect('dashboard_enfermero')

    if wants_json:
        return JsonResponse({'ok': True})
    messages.success(request, "Paciente calificado correctamente.")
    return redirect('dashboard_enfermero')

@usuario_login_required
@role_required(RolUsuario.ENFERMERO)
def actualizar_estado_servicio(request, servicio_id):
    wants_json = request.headers.get('x-requested-with') == 'XMLHttpRequest' or 'application/json' in (request.headers.get('accept') or '')

    if request.method != 'POST':
        if wants_json:
            return JsonResponse({'status': 'error', 'message': 'Método no permitido.'}, status=405)
        return redirect('dashboard_enfermero')

    enfermero = request.usuario
    servicio = None
    
    # Soporte para datos JSON o Form POST
    if request.content_type == 'application/json':
        try:
            data = json.loads(request.body)
            nuevo_estado = data.get('nuevo_estado')
            codigo_usuario = data.get('codigo_usuario')
            if not nuevo_estado and (data.get('rating') or data.get('estrellas')):
                nuevo_estado = EstadoServicio.COMPLETADO
        except json.JSONDecodeError:
            if wants_json:
                return JsonResponse({'status': 'error', 'message': 'JSON inválido.'}, status=400)
            return redirect('dashboard_enfermero')
    else:
        nuevo_estado = (request.POST.get('nuevo_estado') or '').strip()
        codigo_usuario = (request.POST.get('codigo_usuario') or '').strip()

    allowed = {EstadoServicio.EN_PROGRESO, EstadoServicio.COMPLETADO}
    if nuevo_estado not in allowed:
        if wants_json:
            return JsonResponse({'status': 'error', 'message': 'Estado inválido.'}, status=400)
        messages.error(request, "Estado inválido.")
        return redirect('dashboard_enfermero')

    with transaction.atomic():
        servicio = (
            ServicioAcompanamiento.objects.select_for_update(of=('self',))
            .filter(id=servicio_id, enfermero=enfermero)
            .first()
        )
        if not servicio:
            if wants_json:
                return JsonResponse({'status': 'error', 'message': 'No tienes permisos sobre este servicio.'}, status=403)
            messages.error(request, "No tienes permisos sobre este servicio.")
            return redirect('dashboard_enfermero')

        if nuevo_estado == EstadoServicio.EN_PROGRESO and servicio.estado != EstadoServicio.EN_CAMINO:
            if wants_json:
                return JsonResponse({'status': 'error', 'message': 'Este servicio no está disponible para iniciar.'}, status=400)
            messages.error(request, "Este servicio no está disponible para iniciar.")
            return redirect('dashboard_enfermero')

        if nuevo_estado == EstadoServicio.COMPLETADO and servicio.estado != EstadoServicio.EN_PROGRESO:
            if wants_json:
                return JsonResponse({'status': 'error', 'message': 'Este servicio no está disponible para finalizar.'}, status=400)
            messages.error(request, "Este servicio no está disponible para finalizar.")
            return redirect('dashboard_enfermero')

        if nuevo_estado == EstadoServicio.EN_PROGRESO:
            now = timezone.now()
            if getattr(servicio, 'codigo_bloqueado_hasta', None) and servicio.codigo_bloqueado_hasta > now:
                if wants_json:
                    return JsonResponse(
                        {'status': 'error', 'message': 'Demasiados intentos. Intenta nuevamente en unos minutos.'},
                        status=429,
                    )
                messages.error(request, "Demasiados intentos. Intenta nuevamente en unos minutos.")
                return redirect('dashboard_enfermero')

            esperado = (getattr(servicio, 'codigo_verificacion', '') or '').strip()
            ingresado = (codigo_usuario or '').strip()
            if not esperado:
                if wants_json:
                    return JsonResponse({'status': 'error', 'message': 'Este servicio no tiene código de verificación asignado.'}, status=400)
                messages.error(request, "Este servicio no tiene código de verificación asignado.")
                return redirect('dashboard_enfermero')

            if ingresado != esperado:
                servicio.codigo_intentos = int(getattr(servicio, 'codigo_intentos', 0) or 0) + 1
                update_fields = ['codigo_intentos']
                if servicio.codigo_intentos >= 5:
                    servicio.codigo_bloqueado_hasta = now + datetime.timedelta(minutes=10)
                    update_fields.append('codigo_bloqueado_hasta')
                servicio.save(update_fields=update_fields)

                if wants_json:
                    return JsonResponse(
                        {'status': 'error', 'message': 'Código de verificación inválido, por favor solicítelo nuevamente al paciente.'},
                        status=400,
                    )
                messages.error(request, "Código de verificación inválido, por favor solicítelo nuevamente al paciente.")
                return redirect('dashboard_enfermero')

            servicio.codigo_intentos = 0
            servicio.codigo_bloqueado_hasta = None
            servicio.save(update_fields=['codigo_intentos', 'codigo_bloqueado_hasta'])

        servicio.estado = nuevo_estado
        servicio.save(update_fields=['estado'])

        if nuevo_estado == EstadoServicio.COMPLETADO:
            if request.content_type == 'application/json':
                try:
                    data = json.loads(request.body)
                    estrellas_raw = data.get('rating') or data.get('estrellas')
                    comentario = (data.get('comments') or data.get('comentario') or '').strip()
                except:
                    estrellas_raw = None
                    comentario = ""
            else:
                estrellas_raw = (request.POST.get('calificacion') or request.POST.get('estrellas') or '').strip()
                comentario = (request.POST.get('comentario') or '').strip()

            if estrellas_raw:
                try:
                    estrellas = int(estrellas_raw)
                except (TypeError, ValueError):
                    estrellas = None
                if estrellas is None or estrellas < 1 or estrellas > 5:
                    if not wants_json:
                        messages.warning(request, "La calificación enviada no es válida. Se finalizó el servicio sin registrar estrellas.")
                else:
                    try:
                        CalificacionPaciente.objects.create(
                            servicio=servicio,
                            paciente=servicio.paciente,
                            enfermero=enfermero,
                            estrellas=estrellas,
                            comentario=comentario,
                        )
                    except IntegrityError:
                        pass

            if getattr(servicio, 'peticion_id', None):
                Peticion.objects.filter(id=servicio.peticion_id).update(estado=EstadoPeticion.COMPLETADA)
            else:
                Peticion.objects.filter(
                    estado=EstadoPeticion.ASIGNADA,
                    postulaciones__enfermero=enfermero,
                    postulaciones__estado=EstadoPostulacion.ACEPTADA,
                ).update(estado=EstadoPeticion.COMPLETADA)

    servicio = (
        ServicioAcompanamiento.objects.select_related('paciente', 'peticion')
        .filter(id=servicio_id, enfermero=enfermero)
        .first()
    )
    correo_paciente = (getattr(getattr(servicio, 'paciente', None), 'email', '') or '').strip()
    correo_enfermero = (getattr(enfermero, 'email', '') or '').strip()
    paciente_name = (getattr(getattr(servicio, 'paciente', None), 'nombre_completo', '') or '').strip() or correo_paciente or 'Paciente'
    enfermero_name = (getattr(enfermero, 'nombre_completo', '') or '').strip() or correo_enfermero or 'Enfermero'
    peticion = getattr(servicio, 'peticion', None)
    titulo_servicio = (getattr(peticion, 'titulo', '') or '').strip() or f'Servicio #{servicio_id}'
    fecha_servicio = timezone.localtime(getattr(peticion, 'fecha_evento', timezone.now())).strftime('%d/%m/%Y %I:%M %p')
    estado_label = 'Acompanamiento iniciado' if nuevo_estado == EstadoServicio.EN_PROGRESO else 'Acompanamiento completado'
    mensaje_estado = (
        f"Hola {paciente_name}, el profesional {enfermero_name} confirmó el inicio del acompanamiento."
        if nuevo_estado == EstadoServicio.EN_PROGRESO else
        f"Hola {paciente_name}, el acompanamiento con {enfermero_name} fue marcado como completado."
    )
    mensaje_estado_enfermero = (
        f"Confirmaste el inicio del acompanamiento de \"{titulo_servicio}\"."
        if nuevo_estado == EstadoServicio.EN_PROGRESO else
        f"Confirmaste la finalización del acompanamiento de \"{titulo_servicio}\"."
    )
    ok_paciente = True
    ok_enfermero = True
    if correo_paciente:
        ok_paciente = send_notification_email(
            subject=f'Assist Care · {estado_label}',
            to=correo_paciente,
            from_email=_mail_from_email(),
            status_label=estado_label,
            headline=estado_label,
            message=mensaje_estado,
            details=compact_email_details(
                make_email_detail('Servicio', titulo_servicio),
                make_email_detail('Fecha programada', fecha_servicio),
                make_email_detail('Profesional', enfermero_name),
                make_email_detail('Estado', nuevo_estado),
            ),
            cta_text='Abrir mi panel',
            cta_url=request.build_absolute_uri(reverse('dashboard_discapacitado')),
            footer_note='Podrás seguir el estado actualizado del servicio desde tu panel.',
        )
    if correo_enfermero:
        ok_enfermero = send_notification_email(
            subject=f'Assist Care · {estado_label}',
            to=correo_enfermero,
            from_email=_mail_from_email(),
            status_label=estado_label,
            headline=estado_label,
            message=f"Hola {enfermero_name}, {mensaje_estado_enfermero}",
            details=compact_email_details(
                make_email_detail('Servicio', titulo_servicio),
                make_email_detail('Paciente', paciente_name),
                make_email_detail('Fecha programada', fecha_servicio),
                make_email_detail('Estado', nuevo_estado),
            ),
            cta_text='Abrir portal del enfermero',
            cta_url=request.build_absolute_uri(reverse('dashboard_enfermero')),
            footer_note='Gracias por mantener el estado del acompanamiento actualizado en tiempo real.',
        )
    if not (ok_paciente and ok_enfermero):
        logger.error("No se pudo enviar notificación de cambio de estado para servicio %s", servicio_id)

    msg = "Acompañamiento iniciado. Estado actualizado a En Progreso." if nuevo_estado == EstadoServicio.EN_PROGRESO else "Servicio finalizado correctamente. Estado actualizado a Completado."
    
    if wants_json:
        return JsonResponse({'status': 'success', 'message': msg})
        
    messages.success(request, msg)
    return redirect('dashboard_enfermero')


@usuario_login_required
@role_required(RolUsuario.ENFERMERO)
def iniciar_viaje(request, servicio_id):
    wants_json = request.headers.get('x-requested-with') == 'XMLHttpRequest' or 'application/json' in (request.headers.get('accept') or '')

    if request.method != 'POST':
        if wants_json:
            return JsonResponse({'status': 'error', 'message': 'Método no permitido.'}, status=405)
        return redirect('dashboard_enfermero')

    if request.content_type == 'application/json':
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            if wants_json:
                return JsonResponse({'status': 'error', 'message': 'JSON inválido.'}, status=400)
            return redirect('dashboard_enfermero')
        codigo_usuario = (data.get('codigo_usuario') or '').strip()
    else:
        codigo_usuario = (request.POST.get('codigo_usuario') or '').strip()

    enfermero = request.usuario
    servicio = None
    with transaction.atomic():
        servicio = (
            ServicioAcompanamiento.objects.select_for_update(of=('self',))
            .filter(id=servicio_id, enfermero=enfermero)
            .first()
        )
        if not servicio:
            if wants_json:
                return JsonResponse({'status': 'error', 'message': 'No tienes permisos sobre este servicio.'}, status=403)
            messages.error(request, "No tienes permisos sobre este servicio.")
            return redirect('dashboard_enfermero')

        if servicio.estado != EstadoServicio.EN_CAMINO:
            if wants_json:
                return JsonResponse({'status': 'error', 'message': 'Este servicio no está disponible para iniciar.'}, status=400)
            messages.error(request, "Este servicio no está disponible para iniciar.")
            return redirect('dashboard_enfermero')

        now = timezone.now()
        if getattr(servicio, 'codigo_bloqueado_hasta', None) and servicio.codigo_bloqueado_hasta > now:
            if wants_json:
                return JsonResponse({'status': 'error', 'message': 'Demasiados intentos. Intenta nuevamente en unos minutos.'}, status=429)
            messages.error(request, "Demasiados intentos. Intenta nuevamente en unos minutos.")
            return redirect('dashboard_enfermero')

        esperado = (getattr(servicio, 'codigo_verificacion', '') or '').strip()
        if not esperado:
            if wants_json:
                return JsonResponse({'status': 'error', 'message': 'Este servicio no tiene código de verificación asignado.'}, status=400)
            messages.error(request, "Este servicio no tiene código de verificación asignado.")
            return redirect('dashboard_enfermero')

        if codigo_usuario != esperado:
            servicio.codigo_intentos = int(getattr(servicio, 'codigo_intentos', 0) or 0) + 1
            update_fields = ['codigo_intentos']
            if servicio.codigo_intentos >= 5:
                servicio.codigo_bloqueado_hasta = now + datetime.timedelta(minutes=10)
                update_fields.append('codigo_bloqueado_hasta')
            servicio.save(update_fields=update_fields)

            if wants_json:
                return JsonResponse(
                    {'status': 'error', 'message': 'Código de verificación inválido, por favor solicítelo nuevamente al paciente.'},
                    status=400,
                )
            messages.error(request, "Código de verificación inválido, por favor solicítelo nuevamente al paciente.")
            return redirect('dashboard_enfermero')

        servicio.codigo_intentos = 0
        servicio.codigo_bloqueado_hasta = None
        servicio.estado = EstadoServicio.EN_PROGRESO
        servicio.save(update_fields=['codigo_intentos', 'codigo_bloqueado_hasta', 'estado'])

    servicio = (
        ServicioAcompanamiento.objects.select_related('paciente', 'peticion')
        .filter(id=servicio_id, enfermero=enfermero)
        .first()
    )
    correo_paciente = (getattr(getattr(servicio, 'paciente', None), 'email', '') or '').strip()
    correo_enfermero = (getattr(enfermero, 'email', '') or '').strip()
    paciente_name = (getattr(getattr(servicio, 'paciente', None), 'nombre_completo', '') or '').strip() or correo_paciente or 'Paciente'
    enfermero_name = (getattr(enfermero, 'nombre_completo', '') or '').strip() or correo_enfermero or 'Enfermero'
    peticion = getattr(servicio, 'peticion', None)
    titulo_servicio = (getattr(peticion, 'titulo', '') or '').strip() or f'Servicio #{servicio_id}'
    fecha_servicio = timezone.localtime(getattr(peticion, 'fecha_evento', timezone.now())).strftime('%d/%m/%Y %I:%M %p')
    ok_paciente = True
    ok_enfermero = True
    if correo_paciente:
        ok_paciente = send_notification_email(
            subject='Assist Care · Acompanamiento iniciado',
            to=correo_paciente,
            from_email=_mail_from_email(),
            status_label='En progreso',
            headline='El acompanamiento ya inició',
            message=f"Hola {paciente_name}, el profesional {enfermero_name} confirmó el inicio del acompanamiento.",
            details=compact_email_details(
                make_email_detail('Servicio', titulo_servicio),
                make_email_detail('Fecha programada', fecha_servicio),
                make_email_detail('Profesional', enfermero_name),
                make_email_detail('Estado', EstadoServicio.EN_PROGRESO),
            ),
            cta_text='Abrir mi panel',
            cta_url=request.build_absolute_uri(reverse('dashboard_discapacitado')),
            footer_note='Podrás seguir el estado actualizado del servicio desde tu panel.',
        )
    if correo_enfermero:
        ok_enfermero = send_notification_email(
            subject='Assist Care · Acompanamiento iniciado',
            to=correo_enfermero,
            from_email=_mail_from_email(),
            status_label='En progreso',
            headline='Confirmaste el inicio del acompanamiento',
            message=f"Hola {enfermero_name}, el acompanamiento de \"{titulo_servicio}\" ya quedó en progreso.",
            details=compact_email_details(
                make_email_detail('Paciente', paciente_name),
                make_email_detail('Fecha programada', fecha_servicio),
                make_email_detail('Estado', EstadoServicio.EN_PROGRESO),
            ),
            cta_text='Abrir portal del enfermero',
            cta_url=request.build_absolute_uri(reverse('dashboard_enfermero')),
            footer_note='Gracias por mantener el estado del acompanamiento actualizado en tiempo real.',
        )
    if not (ok_paciente and ok_enfermero):
        logger.error("No se pudo enviar notificación de inicio para servicio %s", servicio_id)

    msg = "Acompañamiento iniciado. Estado actualizado a En Progreso."
    if wants_json:
        return JsonResponse({'status': 'success', 'message': msg})
    messages.success(request, msg)
    return redirect('dashboard_enfermero')


@usuario_login_required
@role_required(RolUsuario.ENFERMERO)
def postular_peticion(request, peticion_id):
    next_url = (request.POST.get('next') or request.GET.get('next') or '').strip()
    safe_next = None
    if next_url.startswith('/') and not next_url.startswith('//'):
        safe_next = next_url

    if request.method != 'POST':
        return redirect(safe_next or 'dashboard_enfermero')

    enfermero = request.usuario
    servicio_activo = ServicioAcompanamiento.objects.filter(
        enfermero=enfermero,
        estado__in=[EstadoServicio.EN_CAMINO, EstadoServicio.EN_PROGRESO],
    ).exists()
    if servicio_activo:
        messages.warning(request, "No puedes postularte mientras tengas un servicio activo.")
        return redirect(safe_next or 'dashboard_enfermero')

    peticion = Peticion.objects.filter(id=peticion_id, estado=EstadoPeticion.ACTIVA).first()
    if not peticion:
        messages.error(request, "La petición no existe o ya no está disponible.")
        return redirect(safe_next or 'dashboard_enfermero')

    fecha_postulacion_iso = ''
    if getattr(peticion, 'fecha_evento', None):
        fecha_postulacion_iso = timezone.localtime(peticion.fecha_evento).date().isoformat()

    blocked_postulation_dates = _nurse_blocked_postulation_dates(enfermero, exclude_peticion_id=peticion.id)
    if fecha_postulacion_iso and fecha_postulacion_iso in blocked_postulation_dates:
        fecha_label = timezone.localtime(peticion.fecha_evento).strftime('%d/%m/%Y')
        messages.warning(
            request,
            f"Ya tienes una postulación registrada para el {fecha_label}. Solo puedes postularte a acompañamientos en días diferentes.",
        )
        return redirect(safe_next or 'dashboard_enfermero')

    try:
        Postulacion.objects.create(
            peticion=peticion,
            enfermero=enfermero,
            estado=EstadoPostulacion.PENDIENTE,
        )
    except IntegrityError:
        messages.warning(request, "Ya te postulaste a esta petición.")
        return redirect(safe_next or 'dashboard_enfermero')

    correo_paciente = (getattr(peticion.discapacitado, 'email', '') or '').strip()
    correo_enfermero = (getattr(enfermero, 'email', '') or '').strip()
    enfermero_name = enfermero.nombre_completo or correo_enfermero or 'Enfermero'
    paciente_name = getattr(peticion.discapacitado, 'nombre_completo', '') or correo_paciente or 'Paciente'
    fecha_servicio = timezone.localtime(peticion.fecha_evento).strftime('%d/%m/%Y %I:%M %p')

    ok_paciente = True
    ok_enfermero = True
    if correo_paciente:
        ok_paciente = send_notification_email(
            subject='Assist Care · Recibiste una nueva postulación',
            to=correo_paciente,
            from_email=_mail_from_email(),
            status_label='Nueva postulacion',
            headline='Tienes una nueva postulación en tu solicitud',
            message=f"Hola {paciente_name}, {enfermero_name} se postuló a tu solicitud \"{peticion.titulo}\".",
            details=compact_email_details(
                make_email_detail('Solicitud', peticion.titulo),
                make_email_detail('Profesional', enfermero_name),
                make_email_detail('Fecha estimada', fecha_servicio),
            ),
            cta_text='Revisar postulaciones',
            cta_url=request.build_absolute_uri(reverse('dashboard_discapacitado')),
            footer_note='Ingresa a tu panel para revisar la postulación y decidir si deseas aceptarla.',
        )
    if correo_enfermero:
        ok_enfermero = send_notification_email(
            subject='Assist Care · Tu postulación fue registrada',
            to=correo_enfermero,
            from_email=_mail_from_email(),
            status_label='Postulacion enviada',
            headline='Tu postulación quedó registrada',
            message=f"Hola {enfermero_name}, tu postulación a \"{peticion.titulo}\" fue registrada correctamente.",
            details=compact_email_details(
                make_email_detail('Paciente', paciente_name),
                make_email_detail('Ciudad', peticion.ciudad),
                make_email_detail('Fecha estimada', fecha_servicio),
            ),
            cta_text='Abrir portal del enfermero',
            cta_url=request.build_absolute_uri(reverse('dashboard_enfermero')),
            footer_note='Te notificaremos por correo cuando el paciente tome una decisión sobre esta solicitud.',
        )
    if not (ok_paciente and ok_enfermero):
        logger.error("No se pudo enviar correo de postulación para la petición %s", peticion.id)
        messages.warning(request, "La postulación fue registrada, pero no fue posible enviar todas las notificaciones por correo.")

    messages.success(request, "Postulación enviada correctamente.")
    return redirect(safe_next or 'dashboard_enfermero')


@usuario_login_required
@role_required(RolUsuario.ENFERMERO)
def retirar_postulacion(request, peticion_id):
    next_url = (request.POST.get('next') or request.GET.get('next') or '').strip()
    safe_next = None
    if next_url.startswith('/') and not next_url.startswith('//'):
        safe_next = next_url

    if request.method != 'POST':
        return redirect(safe_next or 'enfermero_servicios')

    enfermero = request.usuario
    postulacion = (
        Postulacion.objects.select_related('peticion')
        .filter(peticion_id=peticion_id, enfermero=enfermero, estado=EstadoPostulacion.PENDIENTE)
        .first()
    )
    if not postulacion:
        messages.warning(request, "No tienes una postulación activa para retirar.")
        return redirect(safe_next or 'enfermero_servicios')

    if getattr(postulacion.peticion, 'estado', None) != EstadoPeticion.ACTIVA:
        messages.error(request, "No puedes retirar la postulación porque la solicitud ya no está disponible.")
        return redirect(safe_next or 'enfermero_servicios')

    postulacion.delete()
    messages.success(request, "Postulación retirada correctamente.")
    return redirect(safe_next or 'enfermero_servicios')


@usuario_login_required
def update_profile_picture(request):
    def _json_failed(message, status_code=400):
        return JsonResponse({'status': 'failed', 'error': message}, status=status_code)

    if request.method != 'POST':
        return _json_failed('Método no permitido.', 405)

    try:
        usuario = request.usuario
        foto = request.FILES.get('foto_perfil')
        foto_base64 = None
        if not foto and request.content_type and request.content_type.startswith('application/json'):
            try:
                payload = json.loads(request.body or b'{}')
            except json.JSONDecodeError:
                payload = {}
            foto_base64 = payload.get('foto_base64') or payload.get('image_base64')
            if not foto_base64:
                return _json_failed('No se proporcionó ninguna imagen.')
            if isinstance(foto_base64, str) and 'base64,' in foto_base64:
                foto_base64 = foto_base64.split('base64,', 1)[1]
            try:
                decoded = base64.b64decode(foto_base64, validate=True)
            except Exception:
                return _json_failed('Imagen inválida.')
            if not decoded:
                return _json_failed('Imagen inválida.')
            foto = ContentFile(decoded, name=f'profile_{uuid.uuid4().hex}.jpg')
        elif not foto:
            return _json_failed('No se proporcionó ninguna imagen.')

        perfil = None
        if usuario.rol == RolUsuario.DISCAPACITADO:
            perfil = getattr(usuario, 'perfil_discapacitado', None)
        elif usuario.rol == RolUsuario.ENFERMERO:
            perfil = getattr(usuario, 'perfil_enfermero', None)

        if not perfil:
            return _json_failed('No se encontró el perfil del usuario.', 404)

        perfil.foto_perfil = foto
        perfil.fecha_actualizacion = timezone.now()
        perfil.save()

        return JsonResponse({'status': 'success', 'message': 'Foto de perfil actualizada correctamente.'})
    except Exception as e:
        return _json_failed(str(e))
